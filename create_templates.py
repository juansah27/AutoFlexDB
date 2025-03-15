import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import argparse

def create_template(process_type, output_dir='static/templates', enhanced=True):
    """
    Membuat file template Excel untuk proses yang ditentukan
    
    Args:
        process_type: Jenis proses ('Bundle', 'Supplementary', 'Gift')
        output_dir: Direktori untuk menyimpan file template
        enhanced: Jika True, tambahkan deskripsi kolom dan contoh data
    """
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Tentukan header berdasarkan jenis proses
    if process_type == 'Bundle':
        headers = ['Client', 'Main SKU', 'Component SKU', 'Qty', 'Start Date', 'End Date']
        descriptions = ["Nama Klien", "SKU Utama", "SKU Komponen", "Jumlah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
        example_row = ["ClientA", "MAIN-001", "COMP-001", "2", "2023-01-01", "2023-12-31"]
    elif process_type == 'Supplementary':
        headers = ['Client', 'ItemID', 'Gift SKU', 'Gift Qty', 'Start Date', 'End Date']
        descriptions = ["Nama Klien", "ID Item", "SKU Hadiah", "Jumlah Hadiah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
        example_row = ["ClientB", "ITEM-001", "GIFT-001", "1", "2023-01-01", "2023-12-31"]
    else:  # Gift
        headers = ['Client', 'SKU', 'GiftSKU', 'Qty', 'Start Date', 'End Date']
        descriptions = ["Nama Klien", "SKU Produk", "SKU Hadiah", "Jumlah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
        example_row = ["ClientC", "PROD-001", "GIFT-001", "1", "2023-01-01", "2023-12-31"]
    
    # Buat style untuk header
    header_fill = PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Tambahkan header dengan style
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Set lebar kolom
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
    
    if enhanced:
        # Tambahkan deskripsi dengan style
        desc_font = Font(italic=True, color='808080')
        desc_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for col_idx, desc in enumerate(descriptions, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = desc
            cell.font = desc_font
            cell.alignment = desc_alignment
            cell.border = thin_border
        
        # Tambahkan contoh data
        example_font = Font(color='0070C0')
        example_alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx, example in enumerate(example_row, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = example
            cell.font = example_font
            cell.alignment = example_alignment
            cell.border = thin_border
    
    # Pastikan direktori ada
    os.makedirs(output_dir, exist_ok=True)
    
    # Simpan file
    template_file = os.path.join(output_dir, f"Template_{process_type}.xlsx")
    wb.save(template_file)
    print(f"Template {process_type} berhasil dibuat: {template_file}")
    
    return template_file

if __name__ == "__main__":
    # Setup parser argumen
    parser = argparse.ArgumentParser(description='Buat file template Excel untuk berbagai jenis proses')
    parser.add_argument('--type', choices=['Bundle', 'Supplementary', 'Gift', 'all'], default='all',
                        help='Jenis proses untuk dibuat template-nya')
    parser.add_argument('--output-dir', default='static/templates', 
                        help='Direktori untuk menyimpan file template')
    parser.add_argument('--simple', action='store_true',
                        help='Jika diberikan, buat template sederhana tanpa deskripsi dan contoh')
    
    args = parser.parse_args()
    
    if args.type == 'all':
        # Buat semua jenis template
        for process_type in ['Bundle', 'Supplementary', 'Gift']:
            create_template(process_type, args.output_dir, not args.simple)
    else:
        # Buat hanya jenis template yang diminta
        create_template(args.type, args.output_dir, not args.simple)
    
    print("Proses pembuatan template selesai!") 