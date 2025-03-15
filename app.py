import os
import pandas as pd
import sqlite3
import logging
import json
import re
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from werkzeug.utils import secure_filename
import secrets
import openpyxl
import database_helper as db_helper
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sys
import hashlib
import tkinter as tk
from tkinter import ttk
from cryptography.fernet import Fernet
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl import load_workbook

# ====================== [CONFIGURATION] =======================
# Logging setup
logging.basicConfig(
    filename="app.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Hapus konfigurasi lisensi dan ganti dengan konfigurasi login
# Konfigurasi User
DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD = "admin123"

# Color Scheme
BG_COLOR = "#181818"  # Dark background
TEXT_COLOR = "#FFFFFF"  # White text
CONTAINER_BG = "#333333"  # Dark gray container
ACCENT_COLOR = "#FF8000"  # Orange
DANGER_COLOR = "#E74C3C"  # Red for errors
SUCCESS_COLOR = "#2ECC71"  # Green
BUTTON_HOVER = "#CC6600"  # Darker orange for hover

# Typography
FONT_PRIMARY = ("Segoe UI", 10)
FONT_HEADING = ("Segoe UI Semibold", 16)
FONT_BUTTON = ("Segoe UI", 10, "bold")

# UI Constants
BORDER_RADIUS = 4
INPUT_PADDING = (5, 3)
CONTENT_PADDING = 10

# Mendapatkan direktori kerja saat ini
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(SCRIPT_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, 'output')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Database untuk menyimpan pengguna
USERS_DB_FILE = os.path.join(SCRIPT_DIR, 'users.json')

# Pastikan folder upload dan output ada
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Fungsi untuk mendapatkan path lengkap file database
def get_db_path(db_name="shop_mapping.db"):
    db_dir = os.path.join(SCRIPT_DIR, 'database')
    os.makedirs(db_dir, exist_ok=True)
    db_path = os.path.join(db_dir, db_name)
    
    # Log informasi path database untuk memudahkan pengecekan
    logging.info(f"Path database yang digunakan: {db_path}")
    return db_path

# Fungsi untuk enkripsi password
def hash_password(password):
    # Fungsi untuk mengenkripsi password
    salt = secrets.token_hex(8)
    h = hashlib.sha256()
    h.update((password + salt).encode('utf-8'))
    return f"{h.hexdigest()}:{salt}"

# Fungsi untuk verifikasi password
def verify_password(stored_password, provided_password):
    # Verifikasi password
    if not stored_password or ":" not in stored_password:
        return False
    
    hash_val, salt = stored_password.split(":")
    h = hashlib.sha256()
    h.update((provided_password + salt).encode('utf-8'))
    return h.hexdigest() == hash_val

# Fungsi untuk menyimpan users
def save_users(users):
    try:
        with open(USERS_DB_FILE, 'w', encoding='utf-8') as file:
            json.dump(users, file, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving users: {e}")
        return False

# Fungsi untuk membuat database pengguna jika belum ada
def create_user_db_if_not_exists():
    if not os.path.exists(USERS_DB_FILE):
        # Buat default admin user dan guest user
        default_users = [
            {
                "username": "admin",
                "name": "Administrator",
                "password": hash_password("admin123"),
                "role": "admin",
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            },
            {
                "username": "guest",
                "name": "Guest User",
                "password": hash_password("guest123"),
                "role": "guest",
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        ]
        save_users(default_users)
        print("File database pengguna dibuat dengan user admin dan guest default")

# Fungsi untuk mendapatkan semua users
def get_users():
    if not os.path.exists(USERS_DB_FILE):
        create_user_db_if_not_exists()
    
    try:
        with open(USERS_DB_FILE, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        print(f"Error loading users: {e}")
        return []

# Fungsi untuk mendapatkan user tanpa password
def get_user(username):
    users = get_users()
    for user in users:
        if user['username'] == username:
            # Jangan kembalikan password ke client
            user_copy = user.copy()
            if 'password' in user_copy:
                del user_copy['password']
            return user_copy
    return None

# Fungsi untuk mendapatkan user dengan password (untuk verifikasi)
def get_user_with_password(username):
    users = get_users()
    for user in users:
        if user['username'] == username:
            return user
    return None

# Fungsi untuk menambahkan user baru
def add_user(username, name, password, role="user"):
    users = get_users()
    
    # Cek apakah username sudah ada
    for user in users:
        if user['username'] == username:
            return False, "Username sudah digunakan"
    
    # Tambahkan user baru
    new_user = {
        "username": username,
        "name": name,
        "password": hash_password(password),
        "role": role,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    users.append(new_user)
    if save_users(users):
        # Kembalikan salinan tanpa password
        user_copy = new_user.copy()
        del user_copy['password']
        return True, user_copy
    return False, "Gagal menyimpan user baru"

# Fungsi untuk update user
def update_user(username, data):
    users = get_users()
    success = False
    
    for i, user in enumerate(users):
        if user['username'] == username:
            # Perbarui informasi yang diizinkan
            if 'name' in data:
                users[i]['name'] = data['name']
            
            # Jika ada permintaan password baru
            if 'current_password' in data and 'new_password' in data:
                # Verifikasi password lama
                if verify_password(user['password'], data['current_password']):
                    users[i]['password'] = hash_password(data['new_password'])
                else:
                    return False, "Password saat ini tidak valid"
            
            success = save_users(users)
            break
    
    if success:
        return True, "Profil berhasil diperbarui"
    return False, "User tidak ditemukan atau gagal memperbarui"

# Fungsi untuk menghapus user
def delete_user(username):
    if username == "admin":
        return False, "User admin tidak bisa dihapus"
    
    users = get_users()
    initial_count = len(users)
    
    users = [user for user in users if user['username'] != username]
    
    if len(users) < initial_count:
        if save_users(users):
            return True, "User berhasil dihapus"
        return False, "Gagal menyimpan perubahan"
    
    return False, "User tidak ditemukan"

# Inisialisasi database SQLite
def initialize_database(db_file=get_db_path()):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS shop_mapping (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    marketplace TEXT NOT NULL,
    client TEXT NOT NULL,
    shop_id TEXT NOT NULL
    )
    """)
    conn.commit()
    conn.close()

# Fungsi untuk membaca mapping dari database
def load_shop_id_mapping_from_db(db_file=get_db_path()):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute("SELECT marketplace, client, shop_id FROM shop_mapping")
    mapping = {f"{str(row[0] or '').lower()}_{str(row[1] or '').lower()}": row[2] for row in cursor.fetchall()}
    conn.close()
    return mapping

# Fungsi untuk menentukan shop_id berdasarkan marketplace dan client
def assign_shop_id(marketplace, client):
    # Pastikan marketplace dan client adalah string
    if isinstance(marketplace, tuple):
        marketplace = ' '.join(marketplace).strip()  # Gabungkan elemen tuple
    if isinstance(client, tuple):
        client = ' '.join(client).strip()  # Gabungkan elemen tuple
        
    # Pastikan marketplace dan client bukan None
    if marketplace is None:
        marketplace = ""
    if client is None:
        client = ""

    # Menangani alias marketplace
    marketplace_mapping = {
        'tokped': 'tokopedia',  # Menambahkan pemetaan untuk tokped
        # Tambahkan pemetaan lain jika diperlukan
    }
    
    # Mengubah marketplace menjadi bentuk standar
    marketplace = marketplace_mapping.get(marketplace.lower().strip(), marketplace.lower().strip())

    key = f"{marketplace}_{client.lower().strip()}"
    logging.info(f"Generated key: {key}")
    shop_id_mapping = load_shop_id_mapping_from_db()
    logging.info(f"Loaded shop ID mapping: {shop_id_mapping}")
    shop_id = shop_id_mapping.get(key, "Unknown")
    if shop_id == "Unknown":
        logging.warning(f"ShopId tidak ditemukan untuk key: {key}")
    logging.info(f"Assigned ShopId: {shop_id}")
    return shop_id

# Fungsi untuk mengubah format tanggal
def convert_date_format(date_value, is_end_date=False):
    if pd.isna(date_value) or date_value == "":
        return "Format Tanggal Salah"
    # Menghapus tanda kurung tetapi mempertahankan waktu
    date_value = re.sub(r'\((\d{2}:\d{2})\)', r'\1:00', date_value).strip()
    possible_formats = [
        "%d/%m/%Y %H.%M", "%d/%m/%Y %H.%M.%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", 
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", 
        "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
        "%m-%d-%Y %H:%M:%S", "%m-%d-%Y %H:%M", "%d/%m/%Y", 
        "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", 
        "%A, %B %d, %Y"
    ]
    for fmt in possible_formats:
        try:
            parsed_date = datetime.strptime(date_value, fmt)
            if is_end_date:
                # Set time to 23:59:59 if only date is provided or time is 00:00:00 or 00:00
                if parsed_date.time() == datetime.min.time():
                    parsed_date = parsed_date.replace(hour=23, minute=59, second=59)
            return parsed_date.strftime("%m/%d/%Y %H:%M:%S")
        except ValueError:
            continue
    return "Format Tanggal Salah"

# Fungsi untuk mendapatkan client_id dari shop_id
def get_client_id(shop_id, db_file=get_db_path()):
    try:
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT client_id FROM shop_mapping WHERE shop_id = ?", (shop_id,))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else None
    except Exception as e:
        logging.error(f"Error getting client_id: {str(e)}")
        return None

# Fungsi untuk mengecek apakah file diizinkan
def allowed_file(filename):
    if filename is None:
        return False
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Middlewares untuk autentikasi
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        
        user = get_user(session['username'])
        if not user or user['role'] != 'admin':
            return jsonify({"success": False, "error": "Akses ditolak. Hanya admin yang diizinkan."}), 403
        
        return f(*args, **kwargs)
    return decorated_function

def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'username' not in session:
                return redirect(url_for('login'))
            
            user = get_user(session['username'])
            if not user or user['role'] not in allowed_roles:
                return jsonify({"success": False, "error": f"Akses ditolak. Anda tidak memiliki izin untuk mengakses fitur ini."}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Inisialisasi aplikasi Flask
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'all-in-one-setting-super-secret-key')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload

# Konfigurasi untuk menghilangkan peringatan keamanan
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['PREFERRED_URL_SCHEME'] = 'http'

# Menambahkan after_request handler untuk mengatur header keamanan global
@app.after_request
def set_secure_headers(response):
    # Menonaktifkan X-Frame-Options untuk mencegah frame blocking
    response.headers['X-Frame-Options'] = 'ALLOWALL'
    # Menonaktifkan Content-Security-Policy yang ketat
    response.headers['Content-Security-Policy'] = "default-src * 'unsafe-inline' 'unsafe-eval'; connect-src * 'unsafe-inline'; img-src * data: blob: 'unsafe-inline'; frame-src *; style-src * 'unsafe-inline'; worker-src * blob: 'unsafe-inline'; frame-ancestors *;"
    # Menghapus Strict-Transport-Security
    if 'Strict-Transport-Security' in response.headers:
        del response.headers['Strict-Transport-Security']
    # Mengizinkan mixed content
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response

# Inisialisasi database
initialize_database()

# Pastikan database pengguna ada saat aplikasi dimulai
create_user_db_if_not_exists()

# =========================== ROUTES ===========================

# Rute untuk halaman utama
@app.route('/')
@login_required
def index():
    return render_template('index.html')

# Rute untuk halaman login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = get_user_with_password(username)
        if user and verify_password(user['password'], password):
            session['username'] = username
            session['role'] = user['role']
            return redirect(url_for('index'))
        
        return render_template('login.html', error='Username atau password salah')
    
    # Jika sudah login, redirect ke homepage
    if 'username' in session:
        return redirect(url_for('index'))
    
    return render_template('login.html')

# Rute untuk logout
@app.route('/logout')
def logout():
    # Log logout
    if 'username' in session:
        logging.info(f"User {session['username']} logout")
    
    # Hapus semua data session
    session.clear()
    
    # Redirect ke halaman login
    return redirect(url_for('login'))

# Rute untuk memproses file yang diunggah
@app.route('/process', methods=['POST'])
@login_required
def process_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang diunggah'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Tidak ada file yang dipilih'})
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Simpan nama file input untuk digunakan di save_output
        app.config['last_input_file'] = filename
        
        process_type = request.form.get('process_type', 'Bundle')
        created_by = request.form.get('created_by', '')
        output_format = request.form.get('output_format', 'xlsx')
        
        try:
            # Memproses file
            with pd.ExcelFile(file_path) as xls:
                all_sheets_data = {}
                client_name = None
                log_messages = []
                
                for sheet_name in xls.sheet_names:
                    # Cek apakah sheet disembunyikan
                    sheet = xls.book[sheet_name]
                    if sheet.sheet_state == 'hidden':
                        log_messages.append(f"Sheet '{sheet_name}' disembunyikan, melewati pemrosesan.")
                        continue
                    
                    log_messages.append(f"Memproses sheet: {sheet_name}")
                    
                    # Membaca beberapa baris pertama untuk menentukan jumlah baris header
                    sample_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=2)
                    
                    # Logika untuk menentukan jumlah baris header
                    header_rows = 0
                    for index, row in sample_df.iterrows():
                        if row.isnull().all():  # Jika seluruh baris kosong, berhenti
                            break
                        header_rows += 1
                    
                    # Membaca DataFrame dengan jumlah baris header yang ditemukan
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=list(range(header_rows)), dtype=str)
                    
                    # Tentukan jenis proses berdasarkan pilihan pengguna
                    if process_type == "Bundle":
                        processed_df, client_name = process_bundle(df, created_by)
                    elif process_type == "Supplementary":
                        processed_df, client_name = process_supplementary(df, created_by)
                    elif process_type == "Gift":
                        processed_df, client_name = process_gift(df, created_by)
                    
                    # Simpan hasil proses
                    all_sheets_data[sheet_name] = processed_df
                
                # Simpan hasil ke file
                output_file = save_output(all_sheets_data, client_name, created_by)
                
                return jsonify({
                    'success': True,
                    'message': 'File berhasil diproses dengan validasi otomatis',
                    'log': log_messages,
                    'output_file': output_file
                })
                
        except Exception as e:
            logging.error(f"Kesalahan saat memproses file: {str(e)}")
            return jsonify({'error': f'Kesalahan saat memproses file: {str(e)}'})
    
    return jsonify({'error': 'Format file tidak didukung'})

# Rute untuk mengunduh file hasil
@app.route('/download/<filename>')
@login_required
def download_file(filename):
    # Buat path file
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    
    # Cek apakah file ada
    if not os.path.exists(file_path):
        return jsonify({'error': 'File tidak ditemukan'}), 404
        
    try:
        # Catat aktivitas download
        logging.info(f"User {session.get('username', 'unknown')} mendownload file {filename}")
        
        # Tambahkan header untuk mengizinkan konten diunduh tanpa peringatan keamanan
        response = send_file(file_path, as_attachment=True)
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['Content-Security-Policy'] = "default-src 'self'"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logging.error(f"Error saat download file: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Endpoint baru untuk download template
@app.route('/download_template/<process_type>')
@login_required
def download_template(process_type):
    try:
        template_file = None
        
        # Tentukan nama file template berdasarkan jenis proses
        if process_type == "Bundle":
            template_file = "Template_Bundle.xlsx"
        elif process_type == "Supplementary":
            template_file = "Template_Supplementary.xlsx"
        elif process_type == "Gift":
            template_file = "Template_Gift.xlsx"
        else:
            return jsonify({'error': 'Jenis proses tidak valid'}), 400
            
        # Buat path file template
        template_path = os.path.join('static', 'templates', template_file)
        
        # Cek apakah file template ada
        if not os.path.exists(template_path):
            # Buat direktori jika belum ada
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            # Jika file tidak ada, buat template baru
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Tambahkan header berdasarkan jenis proses
            if process_type == "Bundle":
                headers = ["Client", "Main SKU", "Component SKU", "Qty", "Start Date", "End Date"]
                # Tambahkan deskripsi kolom pada baris kedua
                descriptions = ["Nama Klien", "SKU Utama", "SKU Komponen", "Jumlah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
            elif process_type == "Supplementary":
                headers = ["Client", "ItemID", "Gift SKU", "Gift Qty", "Start Date", "End Date"]
                # Tambahkan deskripsi kolom pada baris kedua
                descriptions = ["Nama Klien", "ID Item", "SKU Hadiah", "Jumlah Hadiah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
            elif process_type == "Gift":
                headers = ["Client", "SKU", "GiftSKU", "Qty", "Start Date", "End Date"]
                # Tambahkan deskripsi kolom pada baris kedua
                descriptions = ["Nama Klien", "SKU Produk", "SKU Hadiah", "Jumlah", "Format: YYYY-MM-DD", "Format: YYYY-MM-DD"]
                
            # Tambahkan kolom contoh untuk format
            example_row = None
            if process_type == "Bundle":
                example_row = ["ClientA", "MAIN-001", "COMP-001", "2", "2023-01-01", "2023-12-31"]
            elif process_type == "Supplementary":
                example_row = ["ClientB", "ITEM-001", "GIFT-001", "1", "2023-01-01", "2023-12-31"]
            elif process_type == "Gift":
                example_row = ["ClientC", "PROD-001", "GIFT-001", "1", "2023-01-01", "2023-12-31"]
                
            # Isi header ke worksheet
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                
                # Tambahkan deskripsi
                desc_cell = ws.cell(row=2, column=col_idx)
                desc_cell.value = descriptions[col_idx-1]
                desc_cell.font = Font(italic=True)
                
                # Tambahkan contoh data
                if example_row:
                    example_cell = ws.cell(row=3, column=col_idx)
                    example_cell.value = example_row[col_idx-1]
                
            # Sesuaikan lebar kolom
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 15
                
            # Simpan file template
            wb.save(template_path)
            
        # Catat aktivitas download template
        logging.info(f"User {session.get('username', 'unknown')} mendownload template {template_file}")
        
        # Set cache headers
        response = send_file(template_path, as_attachment=True, download_name=template_file)
        response.headers['Cache-Control'] = 'public, max-age=86400'  # Cache selama 1 hari
        response.headers['X-Content-Type-Options'] = 'nosniff'
        return response
        
    except Exception as e:
        logging.error(f"Error saat download template: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Rute untuk manajemen database
@app.route('/database', methods=['GET'])
@role_required(['admin'])
def database_view():
    """Tampilkan halaman manajemen database"""
    return render_template('database.html')

@app.route('/api/shop_mappings', methods=['GET'])
def get_shop_mappings():
    """API endpoint untuk mendapatkan semua mapping toko"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute("SELECT id, marketplace, client, shop_id FROM shop_mapping ORDER BY marketplace, client")
        mappings = [
            {
                'id': row[0], 
                'marketplace': row[1], 
                'client': row[2], 
                'shop_id': row[3]
            } 
            for row in cursor.fetchall()
        ]
        conn.close()
        return jsonify({'success': True, 'data': mappings})
    except Exception as e:
        logging.error(f"Kesalahan saat mengambil data mapping: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop_mappings', methods=['POST'])
@login_required
def add_shop_mapping():
    """API endpoint untuk menambahkan mapping baru"""
    try:
        data = request.get_json()
        marketplace = data.get('marketplace')
        client = data.get('client')
        shop_id = data.get('shop_id')
        
        if not marketplace or not client or not shop_id:
            return jsonify({'success': False, 'error': 'Semua field (marketplace, client, shop_id) wajib diisi'})
        
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        
        # Cek apakah kombinasi marketplace dan client sudah ada
        cursor.execute(
            "SELECT id FROM shop_mapping WHERE marketplace = ? AND client = ?",
            (marketplace, client)
        )
        existing = cursor.fetchone()
        
        if existing:
            # Update jika sudah ada
            cursor.execute(
                "UPDATE shop_mapping SET shop_id = ? WHERE marketplace = ? AND client = ?",
                (shop_id, marketplace, client)
            )
            message = "Data berhasil diperbarui"
        else:
            # Insert jika belum ada
            cursor.execute(
                "INSERT INTO shop_mapping (marketplace, client, shop_id) VALUES (?, ?, ?)",
                (marketplace, client, shop_id)
            )
            message = "Data berhasil ditambahkan"
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        logging.error(f"Kesalahan saat menambahkan data mapping: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/shop_mappings/<int:mapping_id>', methods=['DELETE'])
@login_required
def delete_shop_mapping(mapping_id):
    """API endpoint untuk menghapus mapping"""
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute("DELETE FROM shop_mapping WHERE id = ?", (mapping_id,))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Data berhasil dihapus'})
        else:
            conn.close()
            return jsonify({'success': False, 'error': 'Data tidak ditemukan'})
    except Exception as e:
        logging.error(f"Kesalahan saat menghapus data mapping: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/import_shop_mappings', methods=['POST'])
@login_required
def import_shop_mappings():
    """API endpoint untuk mengimpor data mapping dari file Excel"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Tidak ada file yang diunggah'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Tidak ada file yang dipilih'})
    
    if file and allowed_file(file.filename):
        try:
            # Simpan file sementara
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Baca file Excel
            df = pd.read_excel(file_path, dtype=str)
            
            # Pastikan kolom yang diperlukan ada
            required_columns = ['marketplace', 'client', 'shop_id']
            column_mapping = {
                'marketplace': 'marketplace',
                'client': 'client',
                'shop id': 'shop_id',
                'shop_id': 'shop_id',
                'toko': 'shop_id'
            }
            
            # Standarisasi nama kolom
            df.columns = [col.lower() for col in df.columns]
            df = df.rename(columns=column_mapping)
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({'success': False, 'error': f'Kolom wajib tidak ditemukan: {", ".join(missing_columns)}'})
            
            # Proses data
            conn = sqlite3.connect(get_db_path())
            cursor = conn.cursor()
            
            count_added = 0
            count_updated = 0
            
            for _, row in df.iterrows():
                marketplace = str(row['marketplace']).strip()
                client = str(row['client']).strip()
                shop_id = str(row['shop_id']).strip()
                
                if not marketplace or not client or not shop_id:
                    continue  # Lewati baris kosong
                
                # Cek apakah kombinasi marketplace dan client sudah ada
                cursor.execute(
                    "SELECT id FROM shop_mapping WHERE marketplace = ? AND client = ?",
                    (marketplace, client)
                )
                existing = cursor.fetchone()
                
                if existing:
                    # Update jika sudah ada
                    cursor.execute(
                        "UPDATE shop_mapping SET shop_id = ? WHERE marketplace = ? AND client = ?",
                        (shop_id, marketplace, client)
                    )
                    count_updated += 1
                else:
                    # Insert jika belum ada
                    cursor.execute(
                        "INSERT INTO shop_mapping (marketplace, client, shop_id) VALUES (?, ?, ?)",
                        (marketplace, client, shop_id)
                    )
                    count_added += 1
            
            conn.commit()
            conn.close()
            
            # Hapus file sementara
            os.remove(file_path)
            
            return jsonify({
                'success': True, 
                'message': f'Berhasil mengimpor data: {count_added} ditambahkan, {count_updated} diperbarui'
            })
            
        except Exception as e:
            logging.error(f"Kesalahan saat mengimpor data mapping: {str(e)}")
            return jsonify({'success': False, 'error': str(e)})
    
    return jsonify({'success': False, 'error': 'Format file tidak didukung'})

# Endpoint untuk riwayat pemrosesan
@app.route('/process_history', methods=['GET'])
@role_required(['admin'])
def process_history_view():
    """Tampilkan halaman riwayat pemrosesan"""
    return render_template('process_history.html')

@app.route('/api/process_history', methods=['GET'])
@role_required(['admin'])
def get_process_history():
    """API endpoint untuk mendapatkan riwayat pemrosesan"""
    try:
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        search = request.args.get('search', '')
        
        if search:
            result = db_helper.search_process_history(search, limit, offset)
        else:
            result = db_helper.get_all_process_history(limit, offset)
        
        return jsonify({
            'success': True,
            'data': result['data'],
            'total': result['total'],
            'limit': result['limit'],
            'offset': result['offset']
        })
    except Exception as e:
        logging.error(f"Kesalahan saat mengambil riwayat pemrosesan: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/process_history/<int:history_id>', methods=['DELETE'])
@role_required(['admin'])
def delete_process_history(history_id):
    """API endpoint untuk menghapus riwayat pemrosesan"""
    try:
        success = db_helper.delete_process_history(history_id)
        if success:
            return jsonify({'success': True, 'message': 'Data riwayat berhasil dihapus'})
        else:
            return jsonify({'success': False, 'error': 'Data riwayat tidak ditemukan'})
    except Exception as e:
        logging.error(f"Kesalahan saat menghapus riwayat pemrosesan: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

# Endpoint untuk user management
@app.route('/users', methods=['GET'])
@role_required(['admin'])
def users_view():
    """Tampilkan halaman manajemen pengguna"""
    return render_template('users.html')

@app.route('/api/users', methods=['GET', 'POST'])
@role_required(['admin'])
def api_users():
    if request.method == 'GET':
        users_list = get_users()
        # Hapus password dari respons
        for user in users_list:
            if 'password' in user:
                del user['password']
        
        return jsonify({"success": True, "data": users_list})
    
    elif request.method == 'POST':
        data = request.json
        if not all(key in data for key in ['username', 'name', 'password']):
            return jsonify({"success": False, "error": "Data tidak lengkap"}), 400
        
        success, result = add_user(
            data['username'], 
            data['name'], 
            data['password'], 
            data.get('role', 'user')
        )
        
        if success:
            return jsonify({"success": True, "message": "User berhasil ditambahkan", "data": result})
        return jsonify({"success": False, "error": result})

@app.route('/api/users/<username>', methods=['DELETE'])
@admin_required
def api_delete_user(username):
    # Pastikan admin tidak bisa menghapus dirinya sendiri
    if username == session['username']:
        return jsonify({"success": False, "error": "Anda tidak dapat menghapus akun Anda sendiri"})
    
    success, message = delete_user(username)
    
    if success:
        return jsonify({"success": True, "message": message})
    return jsonify({"success": False, "error": message})

@app.route('/profile', methods=['GET'])
@login_required
def profile_view():
    """Tampilkan halaman profil pengguna"""
    return render_template('profile.html')

@app.route('/api/profile', methods=['GET', 'POST'])
@login_required
def api_profile():
    if request.method == 'GET':
        user = get_user(session['username'])
        if user:
            return jsonify({"success": True, "data": user})
        return jsonify({"success": False, "error": "User tidak ditemukan"})
    
    elif request.method == 'POST':
        data = request.json
        success, message = update_user(session['username'], data)
        
        if success:
            return jsonify({"success": True, "message": message})
        return jsonify({"success": False, "error": message})

# Fungsi untuk memproses Bundle
def process_bundle(df, created_by="System"):
    try:
        # Kolom yang diperlukan
        kolom_wajib = ["SKU Bundle", "SKU Product", "Qty", "Client", "Marketplace", "Start_Date", "End_Date"]
        
        # Debug: Cetak nama kolom sebelum pemrosesan
        print("Kolom asli:", df.columns.tolist())
        
        # Gabungkan header dan buat nama kolom yang sesuai
        df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
        
        # Debug: Cetak nama kolom setelah penggabungan
        print("Kolom setelah penggabungan:", df.columns.tolist())
        
        # Mapping nama kolom yang sesuai
        column_mapping = {
            'Client Unnamed: 1_level_1': 'Client',
            'Client Unnamed: 2_level_1': 'Client',
            'Client Unnamed: 3_level_1': 'Client',
            'Client Unnamed: 4_level_1': 'Client',
            'Client Unnamed: 5_level_1': 'Client',
            'Client Unnamed: 6_level_1': 'Client',
            'Client Unnamed: 7_level_1': 'Client',
            'Client Unnamed: 8_level_1': 'Client',
            'Client Unnamed: 9_level_1': 'Client',
            'Client Unnamed: 10_level_1': 'Client',
            'Client Unnamed: 11_level_1': 'Client',
            'Client Unnamed: 12_level_1': 'Client',
            'Client Unnamed: 13_level_1': 'Client',
            'Client Unnamed: 14_level_1': 'Client',
            'Client Unnamed: 15_level_1': 'Client',
            'Client Unnamed: 16_level_1': 'Client',
            'Client Unnamed: 17_level_1': 'Client',
            'Client Unnamed: 18_level_1': 'Client',
            'Client Unnamed: 19_level_1': 'Client',
            'Client Unnamed: 20_level_1': 'Client',
            'End_Date Unnamed: 1_level_1': 'End_Date',
            'End_Date Unnamed: 2_level_1': 'End_Date',
            'End_Date Unnamed: 3_level_1': 'End_Date',
            'End_Date Unnamed: 4_level_1': 'End_Date',
            'End_Date Unnamed: 5_level_1': 'End_Date',
            'End_Date Unnamed: 6_level_1': 'End_Date',
            'End_Date Unnamed: 7_level_1': 'End_Date',
            'End_Date Unnamed: 8_level_1': 'End_Date',
            'End_Date Unnamed: 9_level_1': 'End_Date',
            'End_Date Unnamed: 10_level_1': 'End_Date',
            'End_Date Unnamed: 11_level_1': 'End_Date',
            'End_Date Unnamed: 12_level_1': 'End_Date',
            'End_Date Unnamed: 13_level_1': 'End_Date',
            'End_Date Unnamed: 14_level_1': 'End_Date',
            'End_Date Unnamed: 15_level_1': 'End_Date',
            'End_Date Unnamed: 16_level_1': 'End_Date',
            'End_Date Unnamed: 17_level_1': 'End_Date',
            'End_Date Unnamed: 18_level_1': 'End_Date',
            'End_Date Unnamed: 19_level_1': 'End_Date',
            'End_Date Unnamed: 20_level_1': 'End_Date',
            'Form Bundle : Client': 'Client',
            'Form Bundle : End_Date': 'End_Date',
            'Form Bundle : Marketplace': 'Marketplace',
            'Form Bundle : SKU BUNDLE': 'SKU Bundle',
            'Form Bundle : SKU Komponen': 'SKU Product',
            'Form Bundle : SKU Komponen.2': 'Qty',
            'Form Bundle : SKU Product': 'SKU Product',
            'Form Bundle : Start_Date': 'Start_Date',
            'Form Bundling List : Client': 'Client',
            'Form Bundling List : End_Date': 'End_Date',
            'Form Bundling List : Marketplace': 'Marketplace',
            'Form Bundling List : SKU Bundling List': 'SKU Bundle',
            'Form Bundling List : SKU Komponen': 'SKU Product',
            'Form Bundling List : SKU Komponen.2': 'Qty',
            'Form Bundling List : Start_Date': 'Start_Date',
            'Market Place Unnamed: 1_level_1': 'Marketplace',
            'Market Place Unnamed: 2_level_1': 'Marketplace',
            'Market Place Unnamed: 3_level_1': 'Marketplace',
            'Market Place Unnamed: 4_level_1': 'Marketplace',
            'Market Place Unnamed: 5_level_1': 'Marketplace',
            'Market Place Unnamed: 6_level_1': 'Marketplace',
            'Market Place Unnamed: 7_level_1': 'Marketplace',
            'Market Place Unnamed: 8_level_1': 'Marketplace',
            'Market Place Unnamed: 9_level_1': 'Marketplace',
            'Market Place Unnamed: 10_level_1': 'Marketplace',
            'Market Place Unnamed: 11_level_1': 'Marketplace',
            'Market Place Unnamed: 12_level_1': 'Marketplace',
            'Market Place Unnamed: 13_level_1': 'Marketplace',
            'Market Place Unnamed: 14_level_1': 'Marketplace',
            'Market Place Unnamed: 15_level_1': 'Marketplace',
            'Market Place Unnamed: 16_level_1': 'Marketplace',
            'Market Place Unnamed: 17_level_1': 'Marketplace',
            'Market Place Unnamed: 18_level_1': 'Marketplace',
            'Market Place Unnamed: 19_level_1': 'Marketplace',
            'Market Place Unnamed: 20_level_1': 'Marketplace',
            'Marketplace Unnamed: 1_level_1': 'Marketplace',
            'Marketplace Unnamed: 2_level_1': 'Marketplace',
            'Marketplace Unnamed: 3_level_1': 'Marketplace',
            'Marketplace Unnamed: 4_level_1': 'Marketplace',
            'Marketplace Unnamed: 5_level_1': 'Marketplace',
            'Marketplace Unnamed: 6_level_1': 'Marketplace',
            'Marketplace Unnamed: 7_level_1': 'Marketplace',
            'Marketplace Unnamed: 8_level_1': 'Marketplace',
            'Marketplace Unnamed: 9_level_1': 'Marketplace',
            'Marketplace Unnamed: 10_level_1': 'Marketplace',
            'Marketplace Unnamed: 11_level_1': 'Marketplace',
            'Marketplace Unnamed: 12_level_1': 'Marketplace',
            'Marketplace Unnamed: 13_level_1': 'Marketplace',
            'Marketplace Unnamed: 14_level_1': 'Marketplace',
            'Marketplace Unnamed: 15_level_1': 'Marketplace',
            'Marketplace Unnamed: 16_level_1': 'Marketplace',
            'Marketplace Unnamed: 17_level_1': 'Marketplace',
            'Marketplace Unnamed: 18_level_1': 'Marketplace',
            'Marketplace Unnamed: 19_level_1': 'Marketplace',
            'Marketplace Unnamed: 20_level_1': 'Marketplace',
            'SKU BUNDLE SKU Bundle': 'SKU Bundle',
            'SKU Bundling List SKU Bundle': 'SKU Bundle',
            'SKU Komponen Qty': 'Qty',
            'SKU Komponen SKU Product': 'SKU Product',
            'Start_Date Unnamed: 1_level_1': 'Start_Date',
            'Start_Date Unnamed: 2_level_1': 'Start_Date',
            'Start_Date Unnamed: 3_level_1': 'Start_Date',
            'Start_Date Unnamed: 4_level_1': 'Start_Date',
            'Start_Date Unnamed: 5_level_1': 'Start_Date',
            'Start_Date Unnamed: 6_level_1': 'Start_Date',
            'Start_Date Unnamed: 7_level_1': 'Start_Date',
            'Start_Date Unnamed: 8_level_1': 'Start_Date',
            'Start_Date Unnamed: 9_level_1': 'Start_Date',
            'Start_Date Unnamed: 10_level_1': 'Start_Date',
            'Start_Date Unnamed: 11_level_1': 'Start_Date',
            'Start_Date Unnamed: 12_level_1': 'Start_Date',
            'Start_Date Unnamed: 13_level_1': 'Start_Date',
            'Start_Date Unnamed: 14_level_1': 'Start_Date',
            'Start_Date Unnamed: 15_level_1': 'Start_Date',
            'Start_Date Unnamed: 16_level_1': 'Start_Date',
            'Start_Date Unnamed: 17_level_1': 'Start_Date',
            'Start_Date Unnamed: 18_level_1': 'Start_Date',
            'Start_Date Unnamed: 19_level_1': 'Start_Date',
            'Start_Date Unnamed: 20_level_1': 'Start_Date'
        }
        
        # Rename kolom
        # Membuat dictionary baru untuk case-insensitive mapping
        lowercase_column_mapping = {}
        for key, value in column_mapping.items():
            lowercase_column_mapping[key.lower()] = value
        
        # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
        rename_dict = {}
        for col in df.columns:
            col_lower = col.lower()
            if col_lower in lowercase_column_mapping:
                rename_dict[col] = lowercase_column_mapping[col_lower]
        
        # Rename kolom menggunakan dictionary yang baru dibuat
        df = df.rename(columns=rename_dict)
        
        # Debug: Cetak nama kolom setelah mapping
        print("Kolom setelah mapping:", df.columns.tolist())
        
        # Cek apakah kolom yang diperlukan ada
        for kolom in kolom_wajib:
            if kolom not in df.columns:
                raise KeyError(f"Kolom '{kolom}' tidak ditemukan")
        
        # Hapus baris yang memiliki nilai kosong pada kolom BOMSKU
        df = df[df["SKU Product"].notna() & (df["SKU Product"] != "")]
        
        # Buat salinan DataFrame
        df = df.copy()
        
        # Fill NaN values
        df.ffill(inplace=True)
        
        # Pisahkan nilai dalam kolom Marketplace dan gandakan baris
        if 'Marketplace' in df.columns:
            df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
            df['Marketplace'] = df['Marketplace'].str.strip()
        
        # Konversi format tanggal
        if 'Start_Date' in df.columns:
            df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
        else:
            raise KeyError("Kolom 'Start_Date' tidak ditemukan")
            
        if 'End_Date' in df.columns:
            df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
        else:
            raise KeyError("Kolom 'End_Date' tidak ditemukan")
        
        # Assign ShopId
        df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
        
        # Hapus kolom Client dari output
        result_df = pd.DataFrame({
            "MainSKU": df["SKU Bundle"],
            "BOMSKU": df["SKU Product"],
            "BOMQty": df["Qty"],
            "IsActive": True,
            "ShopId": df["ShopId"],
            "CreatedDate": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
            "CreatedBy": created_by,
            "UpdatedDate": "NULL",
            "UpdatedBy": "NULL",
            "SingleSKU": False,
            "IsHadiah": False,
            "StartDate": df["Start_Date"],
            "EndDate": df["End_Date"],
            "Multiply": True
        })
        
        # Setelah pemetaan kolom
        df.columns = df.columns.str.replace(' ', '').str.strip()
        
        return result_df, df["Client"].iloc[0]  # Return client name separately
        
    except Exception as e:
        logging.error(f"Error dalam process_bundle: {str(e)}")
        raise

# Fungsi untuk memproses Supplementary
def process_supplementary(df, created_by="System"):
    try:
        # Debug: Cetak nama kolom sebelum pemrosesan
        print("Kolom asli:", df.columns.tolist())
        
        # Gabungkan header dan buat nama kolom yang sesuai
        df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
        
        # Debug: Cetak nama kolom setelah penggabungan
        print("Kolom setelah penggabungan:", df.columns.tolist())
        
        # Mapping nama kolom yang sesuai untuk input
        column_mapping = {
            'Client Unnamed: 1_level_1': 'Client',
            'Client Unnamed: 2_level_1': 'Client',
            'Client Unnamed: 3_level_1': 'Client',
            'Client Unnamed: 4_level_1': 'Client',
            'Client Unnamed: 5_level_1': 'Client',
            'Client Unnamed: 6_level_1': 'Client',
            'Client Unnamed: 7_level_1': 'Client',
            'Client Unnamed: 8_level_1': 'Client',
            'Client Unnamed: 9_level_1': 'Client',
            'Client Unnamed: 10_level_1': 'Client',
            'Client Unnamed: 11_level_1': 'Client',
            'Client Unnamed: 12_level_1': 'Client',
            'Client Unnamed: 13_level_1': 'Client',
            'Client Unnamed: 14_level_1': 'Client',
            'Client Unnamed: 15_level_1': 'Client',
            'Client Unnamed: 16_level_1': 'Client',
            'Client Unnamed: 17_level_1': 'Client',
            'Client Unnamed: 18_level_1': 'Client',
            'Client Unnamed: 19_level_1': 'Client',
            'Client Unnamed: 20_level_1': 'Client',
            'Gift Qty Unnamed: 1_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 2_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 3_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 4_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 5_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 6_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 7_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 8_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 9_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 10_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 11_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 12_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 13_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 14_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 15_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 16_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 17_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 18_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 19_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 20_level_1': 'Gift Qty',
            'Gift SKU Unnamed: 1_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 2_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 3_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 4_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 5_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 6_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 7_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 8_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 9_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 10_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 11_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 12_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 13_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 14_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 15_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 16_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 17_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 18_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 19_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 20_level_1': 'Gift SKU',
            'Main SKU Unnamed: 1_level_1': 'Main SKU',
            'Main SKU Unnamed: 2_level_1': 'Main SKU',
            'Main SKU Unnamed: 3_level_1': 'Main SKU',
            'Main SKU Unnamed: 4_level_1': 'Main SKU',
            'Main SKU Unnamed: 5_level_1': 'Main SKU',
            'Main SKU Unnamed: 6_level_1': 'Main SKU',
            'Main SKU Unnamed: 7_level_1': 'Main SKU',
            'Main SKU Unnamed: 8_level_1': 'Main SKU',
            'Main SKU Unnamed: 9_level_1': 'Main SKU',
            'Main SKU Unnamed: 10_level_1': 'Main SKU',
            'Main SKU Unnamed: 11_level_1': 'Main SKU',
            'Main SKU Unnamed: 12_level_1': 'Main SKU',
            'Main SKU Unnamed: 13_level_1': 'Main SKU',
            'Main SKU Unnamed: 14_level_1': 'Main SKU',
            'Main SKU Unnamed: 15_level_1': 'Main SKU',
            'Main SKU Unnamed: 16_level_1': 'Main SKU',
            'Main SKU Unnamed: 17_level_1': 'Main SKU',
            'Main SKU Unnamed: 18_level_1': 'Main SKU',
            'Main SKU Unnamed: 19_level_1': 'Main SKU',
            'Main SKU Unnamed: 20_level_1': 'Main SKU',
            'Market Place Unnamed: 1_level_1': 'Marketplace',
            'Market Place Unnamed: 2_level_1': 'Marketplace',
            'Market Place Unnamed: 3_level_1': 'Marketplace',
            'Market Place Unnamed: 4_level_1': 'Marketplace',
            'Market Place Unnamed: 5_level_1': 'Marketplace',
            'Market Place Unnamed: 6_level_1': 'Marketplace',
            'Market Place Unnamed: 7_level_1': 'Marketplace',
            'Market Place Unnamed: 8_level_1': 'Marketplace',
            'Market Place Unnamed: 9_level_1': 'Marketplace',
            'Market Place Unnamed: 10_level_1': 'Marketplace',
            'Market Place Unnamed: 11_level_1': 'Marketplace',
            'Market Place Unnamed: 12_level_1': 'Marketplace',
            'Market Place Unnamed: 13_level_1': 'Marketplace',
            'Market Place Unnamed: 14_level_1': 'Marketplace',
            'Market Place Unnamed: 15_level_1': 'Marketplace',
            'Market Place Unnamed: 16_level_1': 'Marketplace',
            'Market Place Unnamed: 17_level_1': 'Marketplace',
            'Market Place Unnamed: 18_level_1': 'Marketplace',
            'Market Place Unnamed: 19_level_1': 'Marketplace',
            'Market Place Unnamed: 20_level_1': 'Marketplace',
            'Marketplace Unnamed: 7_level_1': 'Marketplace',
            'Marketplace Unnamed: 8_level_1': 'Marketplace',
            'Marketplace Unnamed: 9_level_1': 'Marketplace',
            'Marketplace Unnamed: 10_level_1': 'Marketplace',
            'Marketplace Unnamed: 11_level_1': 'Marketplace',
            'Marketplace Unnamed: 12_level_1': 'Marketplace',
            'Marketplace Unnamed: 13_level_1': 'Marketplace',
            'Marketplace Unnamed: 14_level_1': 'Marketplace',
            'Marketplace Unnamed: 15_level_1': 'Marketplace',
            'Marketplace Unnamed: 16_level_1': 'Marketplace',
            'Marketplace Unnamed: 17_level_1': 'Marketplace',
            'Marketplace Unnamed: 18_level_1': 'Marketplace',
            'Marketplace Unnamed: 19_level_1': 'Marketplace',
            'Marketplace Unnamed: 20_level_1': 'Marketplace',
            'Periode End_Date': 'End_Date',
            'End_Date Unnamed: 1_level_1': 'End_Date',
            'End_Date Unnamed: 2_level_1': 'End_Date',
            'End_Date Unnamed: 3_level_1': 'End_Date',
            'End_Date Unnamed: 4_level_1': 'End_Date',
            'End_Date Unnamed: 5_level_1': 'End_Date',
            'End_Date Unnamed: 6_level_1': 'End_Date',
            'End_Date Unnamed: 7_level_1': 'End_Date',
            'End_Date Unnamed: 8_level_1': 'End_Date',
            'End_Date Unnamed: 9_level_1': 'End_Date',
            'End_Date Unnamed: 10_level_1': 'End_Date',
            'End_Date Unnamed: 11_level_1': 'End_Date',
            'End_Date Unnamed: 12_level_1': 'End_Date',
            'End_Date Unnamed: 13_level_1': 'End_Date',
            'End_Date Unnamed: 14_level_1': 'End_Date',
            'End_Date Unnamed: 15_level_1': 'End_Date',
            'End_Date Unnamed: 16_level_1': 'End_Date',
            'End_Date Unnamed: 17_level_1': 'End_Date',
            'End_Date Unnamed: 18_level_1': 'End_Date',
            'End_Date Unnamed: 19_level_1': 'End_Date',
            'End_Date Unnamed: 20_level_1': 'End_Date',
            'Periode Start_Date': 'Start_Date',
            'Start_Date Unnamed: 1_level_1': 'Start_Date',
            'Start_Date Unnamed: 2_level_1': 'Start_Date',
            'Start_Date Unnamed: 3_level_1': 'Start_Date',
            'Start_Date Unnamed: 4_level_1': 'Start_Date',
            'Start_Date Unnamed: 5_level_1': 'Start_Date',
            'Start_Date Unnamed: 6_level_1': 'Start_Date',
            'Start_Date Unnamed: 7_level_1': 'Start_Date',
            'Start_Date Unnamed: 8_level_1': 'Start_Date',
            'Start_Date Unnamed: 9_level_1': 'Start_Date',
            'Start_Date Unnamed: 10_level_1': 'Start_Date',
            'Start_Date Unnamed: 11_level_1': 'Start_Date',
            'Start_Date Unnamed: 12_level_1': 'Start_Date',
            'Start_Date Unnamed: 13_level_1': 'Start_Date',
            'Start_Date Unnamed: 14_level_1': 'Start_Date',
            'Start_Date Unnamed: 15_level_1': 'Start_Date',
            'Start_Date Unnamed: 16_level_1': 'Start_Date',
            'Start_Date Unnamed: 17_level_1': 'Start_Date',
            'Start_Date Unnamed: 18_level_1': 'Start_Date',
            'Start_Date Unnamed: 19_level_1': 'Start_Date',
            'Start_Date Unnamed: 20_level_1': 'Start_Date',
            'Brand Unnamed: 1_level_1': 'Client',
            'Brand Unnamed: 2_level_1': 'Client',
            'Brand Unnamed: 3_level_1': 'Client',
            'Brand Unnamed: 4_level_1': 'Client',
            'Brand Unnamed: 5_level_1': 'Client',
            'Brand Unnamed: 6_level_1': 'Client',
            'Brand Unnamed: 7_level_1': 'Client',
            'Brand Unnamed: 8_level_1': 'Client',
            'Brand Unnamed: 9_level_1': 'Client',
            'Brand Unnamed: 10_level_1': 'Client',
            'Brand Unnamed: 11_level_1': 'Client',
            'Brand Unnamed: 12_level_1': 'Client',
            'Brand Unnamed: 13_level_1': 'Client',
            'Brand Unnamed: 14_level_1': 'Client',
            'Brand Unnamed: 15_level_1': 'Client',
            'Brand Unnamed: 16_level_1': 'Client',
            'Brand Unnamed: 17_level_1': 'Client',
            'Brand Unnamed: 18_level_1': 'Client',
            'Brand Unnamed: 19_level_1': 'Client',
            'Brand Unnamed: 20_level_1': 'Client',
        }
        
        # Rename kolom
        # Membuat dictionary baru untuk case-insensitive mapping
        lowercase_column_mapping = {}
        for key, value in column_mapping.items():
            lowercase_column_mapping[key.lower()] = value
        
        # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
        rename_dict = {}
        for col in df.columns:
            col_lower = col.lower()
            if col_lower in lowercase_column_mapping:
                rename_dict[col] = lowercase_column_mapping[col_lower]
        
        # Rename kolom menggunakan dictionary yang baru dibuat
        df = df.rename(columns=rename_dict)
        df = df.copy()
        df.ffill(inplace=True)
        
        # Debug: Cetak nama kolom setelah mapping
        print("Kolom setelah mapping:", df.columns.tolist())
        
        # Menyeragamkan nama kolom MarketPlace menjadi Marketplace
        if 'MarketPlace' in df.columns and 'Marketplace' not in df.columns:
            df = df.rename(columns={'MarketPlace': 'Marketplace'})
            print("Renamed 'MarketPlace' to 'Marketplace' for consistency")
        
        # Memastikan kolom 'Marketplace' ada
        if 'Marketplace' not in df.columns:
            raise KeyError("Kolom 'Marketplace' tidak ditemukan setelah pemetaan.")
        
        # Memastikan kolom 'Client' ada
        if 'Client' not in df.columns:
            raise KeyError("Kolom 'Client' tidak ditemukan setelah pemetaan.")
        
        # Pisahkan nilai dalam kolom Marketplace
        if 'Marketplace' in df.columns:
            df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
            df['Marketplace'] = df['Marketplace'].str.strip()
        else:
            raise KeyError("Kolom 'Marketplace' tidak ditemukan setelah pemetaan.")
        
        # Konversi format tanggal
        if 'Start_Date' in df.columns:
            df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
        if 'End_Date' in df.columns:
            df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
        
        # Assign ShopId dengan memperhitungkan Client
        if 'Client' not in df.columns:
            raise KeyError("Kolom 'Client' tidak ditemukan setelah pemetaan.")
        
        df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
        
        # Hapus kolom Client dari output
        result_df = pd.DataFrame({
            "ItemID": df["Main SKU"],
            "Supplementary": df["Gift SKU"],
            "SupplementaryQty": df["Gift Qty"],
            "IsActive": True,
            "StartDate": df["Start_Date"],
            "EndDate": df["End_Date"],
            "ShopID": df["ShopId"],
            "CreatedAt": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
            "CreatedBy": created_by
        })
        
        # Setelah pemetaan kolom
        df.columns = df.columns.str.replace(' ', '').str.strip()
        
        return result_df, df["Client"].iloc[0]  # Return client name separately
        
    except Exception as e:
        logging.error(f"Error dalam process_supplementary: {str(e)}")
        raise

def process_gift(df, created_by="System"):
    try:
        # Cek apakah kolom 'Main SKU' ada
        if 'Main SKU' not in df.columns:
            df['Main SKU'] = None  # Atur nilai menjadi NULL jika kolom tidak ada
        
        # Debug: Cetak nama kolom sebelum pemrosesan
        print("Kolom asli:", df.columns.tolist())
        
        # Gabungkan header dan buat nama kolom yang sesuai
        df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
        
        # Debug: Cetak nama kolom setelah penggabungan
        print("Kolom setelah penggabungan:", df.columns.tolist())
        
        # Mapping nama kolom yang sesuai untuk input
        column_mapping = {
            'CLIENT Unnamed: 13_level_1': 'Client',
            'Client Unnamed: 0_level_1': 'Client',
            'Client Unnamed: 1_level_1': 'Client',
            'Client Unnamed: 2_level_1': 'Client',
            'Client Unnamed: 3_level_1': 'Client',
            'Client Unnamed: 4_level_1': 'Client',
            'Client Unnamed: 5_level_1': 'Client',
            'Client Unnamed: 6_level_1': 'Client',
            'Client Unnamed: 7_level_1': 'Client',
            'Client Unnamed: 8_level_1': 'Client',
            'Client Unnamed: 9_level_1': 'Client',
            'Client Unnamed: 10_level_1': 'Client',
            'Client Unnamed: 11_level_1': 'Client',
            'Client Unnamed: 12_level_1': 'Client',
            'Client Unnamed: 13_level_1': 'Client',
            'Client Unnamed: 14_level_1': 'Client',
            'Client Unnamed: 15_level_1': 'Client',
            'Client Unnamed: 16_level_1': 'Client',
            'Client Unnamed: 17_level_1': 'Client',
            'Client Unnamed: 18_level_1': 'Client',
            'Client Unnamed: 19_level_1': 'Client',
            'Client Unnamed: 20_level_1': 'Client',
            'Gift Qty Unnamed: 7_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 8_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 9_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 10_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 11_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 12_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 13_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 14_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 15_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 16_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 17_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 18_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 19_level_1': 'Gift Qty',
            'Gift Qty Unnamed: 20_level_1': 'Gift Qty',
            'Gift SKU Unnamed: 5_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 6_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 7_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 8_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 9_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 10_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 11_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 12_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 13_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 14_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 15_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 16_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 17_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 18_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 19_level_1': 'Gift SKU',
            'Gift SKU Unnamed: 20_level_1': 'Gift SKU',
            'Limit Qty Unnamed: 7_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 8_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 9_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 10_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 11_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 12_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 13_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 14_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 15_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 16_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 17_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 18_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 19_level_1': 'Limit Qty',
            'Limit Qty Unnamed: 20_level_1': 'Limit Qty',
            'Limit Unnamed: 7_level_1': 'Limit Qty',
            'Limit Unnamed: 8_level_1': 'Limit Qty',
            'Limit Unnamed: 9_level_1': 'Limit Qty',
            'Limit Unnamed: 10_level_1': 'Limit Qty',
            'Limit Unnamed: 11_level_1': 'Limit Qty',
            'Limit Unnamed: 12_level_1': 'Limit Qty',
            'Limit Unnamed: 13_level_1': 'Limit Qty',
            'Limit Unnamed: 14_level_1': 'Limit Qty',
            'Limit Unnamed: 15_level_1': 'Limit Qty',
            'Limit Unnamed: 16_level_1': 'Limit Qty',
            'Limit Unnamed: 17_level_1': 'Limit Qty',
            'Limit Unnamed: 18_level_1': 'Limit Qty',
            'Limit Unnamed: 19_level_1': 'Limit Qty',
            'Limit Unnamed: 20_level_1': 'Limit Qty',
            'Main SKU Unnamed: 5_level_1': 'Main SKU',
            'Main SKU Unnamed: 6_level_1': 'Main SKU',
            'Main SKU Unnamed: 7_level_1': 'Main SKU',
            'Main SKU Unnamed: 8_level_1': 'Main SKU',
            'Main SKU Unnamed: 9_level_1': 'Main SKU',
            'Main SKU Unnamed: 10_level_1': 'Main SKU',
            'Main SKU Unnamed: 11_level_1': 'Main SKU',
            'Main SKU Unnamed: 12_level_1': 'Main SKU',
            'Main SKU Unnamed: 13_level_1': 'Main SKU',
            'Main SKU Unnamed: 14_level_1': 'Main SKU',
            'Main SKU Unnamed: 15_level_1': 'Main SKU',
            'Main SKU Unnamed: 16_level_1': 'Main SKU',
            'Main SKU Unnamed: 17_level_1': 'Main SKU',
            'Main SKU Unnamed: 18_level_1': 'Main SKU',
            'Main SKU Unnamed: 19_level_1': 'Main SKU',
            'Main SKU Unnamed: 20_level_1': 'Main SKU',
            'Market Place Unnamed: 8_level_1': 'Marketplace',
            'Market Place Unnamed: 9_level_1': 'Marketplace',
            'Market Place Unnamed: 10_level_1': 'Marketplace',
            'Market Place Unnamed: 11_level_1': 'Marketplace',
            'Market Place Unnamed: 12_level_1': 'Marketplace',
            'Market Place Unnamed: 13_level_1': 'Marketplace',
            'Market Place Unnamed: 14_level_1': 'Marketplace',
            'Market Place Unnamed: 15_level_1': 'Marketplace',
            'Market Place Unnamed: 16_level_1': 'Marketplace',
            'Market Place Unnamed: 17_level_1': 'Marketplace',
            'Market Place Unnamed: 18_level_1': 'Marketplace',
            'Market Place Unnamed: 19_level_1': 'Marketplace',
            'Market Place Unnamed: 20_level_1': 'Marketplace',
            'Marketplace Unnamed: 7_level_1': 'Marketplace',
            'Marketplace Unnamed: 8_level_1': 'Marketplace',
            'Marketplace Unnamed: 9_level_1': 'Marketplace',
            'Marketplace Unnamed: 10_level_1': 'Marketplace',
            'Marketplace Unnamed: 11_level_1': 'Marketplace',
            'Marketplace Unnamed: 12_level_1': 'Marketplace',
            'Marketplace Unnamed: 13_level_1': 'Marketplace',
            'Marketplace Unnamed: 14_level_1': 'Marketplace',
            'Marketplace Unnamed: 15_level_1': 'Marketplace',
            'Marketplace Unnamed: 16_level_1': 'Marketplace',
            'Marketplace Unnamed: 17_level_1': 'Marketplace',
            'Marketplace Unnamed: 18_level_1': 'Marketplace',
            'Marketplace Unnamed: 19_level_1': 'Marketplace',
            'Marketplace Unnamed: 20_level_1': 'Marketplace',
            'Periode End_Date': 'End_Date',
            'Periode Start_Date': 'Start_Date',
            'Purchase (Min -> Max) Value_End': 'Value_End',
            'Purchase (Min -> Max) Value_Start': 'Value_Start'
        }
        
        # Rename kolom
        # Membuat dictionary baru untuk case-insensitive mapping
        lowercase_column_mapping = {}
        for key, value in column_mapping.items():
            lowercase_column_mapping[key.lower()] = value
        
        # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
        rename_dict = {}
        for col in df.columns:
            col_lower = col.lower()
            if col_lower in lowercase_column_mapping:
                rename_dict[col] = lowercase_column_mapping[col_lower]
        
        # Rename kolom menggunakan dictionary yang baru dibuat
        df = df.rename(columns=rename_dict)
        df = df.copy()
        df.ffill(inplace=True)
        
        # Debug: Cetak nama kolom setelah mapping
        print("Kolom setelah mapping:", df.columns.tolist())
        
        # Pisahkan nilai dalam kolom Marketplace
        if 'Marketplace' in df.columns:
            df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
            df['Marketplace'] = df['Marketplace'].str.strip()
        else:
            raise KeyError("Kolom 'Marketplace' tidak ditemukan setelah pemetaan.")
        
        # Ganti nilai "No Limit" dengan "NULL" pada kolom yang relevan
        if 'Limit Qty' in df.columns:
            df['Limit Qty'] = df['Limit Qty'].replace("No Limit", "NULL")
        
        # Assign ShopId dengan memperhitungkan Client
        df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
        
        # Konversi format tanggal
        if 'Start_Date' in df.columns:
            df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
        else:
            print("Kolom yang tersedia:", df.columns.tolist())
            raise KeyError("Kolom 'Start_Date' tidak ditemukan")
            
        if 'End_Date' in df.columns:
            df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
        
        # Generate GIFT ID tanpa groupby (sequential number)
        df['GIFT_ID'] = range(1, len(df) + 1)
        
        # Tambahkan counter untuk setiap kombinasi marketplace dan client
        df['MARKET_COUNTER'] = df.groupby(['Marketplace', 'Client']).cumcount() + 1
        
        # Fungsi untuk mendapatkan kode marketplace
        def get_marketplace_code(marketplace):
            marketplace_codes = {
                'shopee': 'SHP',
                'tiktok': 'TTS',
                'tokopedia': 'TKP',
                'lazada': 'LZD',
                'blibli': 'BLI',
                'zalora': 'ZAL',
                'jubelio': 'JBL',
                'desty': 'DST'
            }
            marketplace = str(marketplace).lower().strip()
            return marketplace_codes.get(marketplace, marketplace[:3].upper())
        
        # Fungsi untuk membuat format GIFT ID
        def create_gift_id(row):
            # Cek value_start untuk kondisi khusus
            value_start = str(row['Value_Start']).replace(',', '').replace('.', '')
            if value_start == "0" or value_start == "":
                value_prefix = "ANY"
            else:
                value_prefix = value_start[:3].zfill(3) + "K"
            
            # Ambil day dari Start_Date dan format lengkap dari End_Date
            try:
                start_date = datetime.strptime(row['Start_Date'], "%m/%d/%Y %H:%M:%S")
                end_date = datetime.strptime(row['End_Date'], "%m/%d/%Y %H:%M:%S")
                
                # Ambil day saja dari start_date
                start_day = f"{start_date.day:02d}"
                # Format lengkap DDMMYY untuk end_date
                end_date_format = f"{end_date.day:02d}{end_date.month:02d}{str(end_date.year)[-2:]}"
            except:
                start_day = "00"
                end_date_format = "000000"
            
            # Proses client prefix
            # Pisahkan string berdasarkan spasi, underscore, titik, atau dash
            client_words = re.split(r'[ _.-]', str(row['Client']).strip())
            client_words = [word for word in client_words if word]  # Hapus empty strings
            
            if len(client_words) >= 3:
                # Jika ada 3 kata atau lebih, ambil huruf pertama dari setiap kata
                client_prefix = (client_words[0][0] + client_words[1][0] + client_words[2][0]).upper()
            elif len(client_words) == 2:
                # Jika ada 2 kata, ambil 3 huruf pertama dari kata pertama
                # dan huruf pertama dari kata kedua
                client_prefix = (client_words[0][:3] + client_words[1][0]).upper()
            else:
                # Jika hanya 1 kata, ambil 3 huruf pertama saja
                client_prefix = client_words[0][:3].upper()
            
            # Dapatkan kode marketplace yang sesuai
            market_prefix = get_marketplace_code(row['Marketplace'])
            
            # Buat format GIFT ID dengan counter di akhir
            gift_id = f"TIER{value_prefix}-{start_day}{end_date_format}-{client_prefix}-{market_prefix}{row['MARKET_COUNTER']}"
            
            return gift_id
        
        # Generate FORMATTED_GIFT_ID
        df['FORMATTED_GIFT_ID'] = df.apply(create_gift_id, axis=1)
        
        # Fungsi untuk menentukan GIFTTYPE
        def determine_gift_type(main_sku):
            return "2" if pd.notna(main_sku) and str(main_sku).strip() != "" else "3"
        
        # Buat DataFrame untuk Header dengan urutan kolom yang ditentukan
        header_columns = [
            "ID", "GIFT ID", "CLIENTID", "SHOPID", "GIFTTYPE", "STARTDATE", 
            "ENDDATE", "ISACTIVE", "CREATED ON", "UPDATE ON", "CREATED BY", 
            "UPDATE DATE", "LIMIT SUMMARY", "USAGE SUMMARY"
        ]
        
        line_columns = [
            "GIFT ID", "MAINSKU", "VALUESTART", "VALUEEND", "GIFTSKU", 
            "GIFTQTY", "GIFTLINENUMBER", "ISACTIVE", "MULTIPLE", 
            "ITEMLIMIT", "ITEMUSAGE"
        ]
        
        header_df = pd.DataFrame({
            "ID": None,
            "GIFT ID": df['FORMATTED_GIFT_ID'],  # Langsung gunakan FORMATTED_GIFT_ID
            "CLIENTID": df['ShopId'].apply(get_client_id),
            "SHOPID": df['ShopId'],
            "GIFTTYPE": df['Main SKU'].apply(determine_gift_type),
            "STARTDATE": df['Start_Date'],
            "ENDDATE": df['End_Date'],
            "ISACTIVE": True,
            "CREATED ON": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
            "UPDATE ON": "NULL",
            "CREATED BY": created_by,
            "UPDATE DATE": "NULL",
            "LIMIT SUMMARY": df.apply(lambda row: row['Limit Qty'] if determine_gift_type(row['Main SKU']) == "2" else None, axis=1),
            "USAGE SUMMARY": "NULL"
        })[header_columns]
        
        # Buat DataFrame untuk Line dengan urutan kolom yang ditentukan
        line_df = pd.DataFrame({
            "GIFT ID": None,
            "MAINSKU": df['Main SKU'],
            "VALUESTART": df['Value_Start'],
            "VALUEEND": df['Value_End'],
            "GIFTSKU": df['Gift SKU'],
            "GIFTQTY": df['Gift Qty'],
            "GIFTLINENUMBER": None,  # Karena tidak ada groupby, setiap baris adalah line pertama
            "ISACTIVE": True,
            "MULTIPLE": False,
            "ITEMLIMIT": df.apply(lambda row: None if determine_gift_type(row['Main SKU']) == "2" else row['Limit Qty'], axis=1),
            "ITEMUSAGE": "NULL"
        })[line_columns]        
        # Setelah pemetaan kolom
        df.columns = df.columns.str.replace(' ', '').str.strip()
        
        return {"Header": header_df, "Line": line_df}, df["Client"].iloc[0]  # Return client name separately
        
    except Exception as e:
        logging.error(f"Error dalam process_gift: {str(e)}")
        raise

# Fungsi untuk menyimpan hasil ke file Excel atau CSV
def save_output(all_sheets_data, client_name, created_by=""):
    try:
        # Ambil nilai dari created_by
        current_date = datetime.today().strftime('%d%m%Y')
        
        # Deteksi process_type berdasarkan key dataframe
        process_type = "Unknown"
        any_df = next(iter(all_sheets_data.values()), None)
        if any_df is not None:
            if isinstance(any_df, dict):
                # Jika ada sub-sheet seperti Header dan Line
                process_type = "Gift"
            elif 'MainSKU' in any_df.columns:
                process_type = "Bundle"
            elif 'ItemID' in any_df.columns:
                process_type = "Supplementary"
        
        # Buat nama file sesuai dengan jenis proses dan ubah menjadi uppercase
        if process_type == "Bundle":
            output_prefix = f"SETUP BUNDLE {client_name} {current_date}".upper()
        elif process_type == "Supplementary":
            output_prefix = f"SETUP SUPPLEMENTARY {client_name} {current_date}".upper()
        elif process_type == "Gift":
            output_prefix = f"SETUP GIFT {client_name} {current_date}".upper()
        else:
            output_prefix = f"SETUP {client_name} {current_date}".upper()
        
        output_file_name = f"{output_prefix}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_file_name)
        
        # Buat workbook baru
        wb = Workbook()
        
        # Mendapatkan file input untuk validasi
        input_file_path = None
        if 'last_input_file' in app.config:
            input_file_path = os.path.join(app.config['UPLOAD_FOLDER'], app.config['last_input_file'])
            logging.info(f"File input untuk validasi: {input_file_path}")
        else:
            logging.warning("Tidak ada file input yang tersimpan di app.config['last_input_file']")
        
        # Simpan setiap DataFrame ke sheet yang berbeda
        for sheet_name, output_df in all_sheets_data.items():
            if isinstance(output_df, dict):
                for sub_sheet_name, df in output_df.items():
                    ws = wb.create_sheet(title=f"{sub_sheet_name} ({len(df)})")
                    # Tulis header
                    headers = df.columns.tolist()
                    for col_idx, header in enumerate(headers, start=1):
                        ws.cell(row=1, column=col_idx, value=header)
                    # Tulis data dengan highlight
                    for row_idx, row in enumerate(df.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            if value == "Unknown":
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Highlight kuning
                            elif value == "Format Tanggal Salah":
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Highlight merah
            else:
                ws = wb.create_sheet(title=f"{sheet_name} ({len(output_df)})")
                # Tulis header
                headers = output_df.columns.tolist()
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                # Tulis data dengan highlight
                for row_idx, row in enumerate(output_df.values, start=2):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if value == "Unknown":
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Highlight kuning
                        elif value == "Format Tanggal Salah":
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Highlight merah
        
        # Hapus sheet default jika ada
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Lakukan validasi antara file input dan output
        validation_results = None
        if input_file_path and os.path.exists(input_file_path):
            try:
                # Buat sheet validasi
                validation_ws = wb.create_sheet(title="Validasi")
                
                # Lakukan validasi dan dapatkan hasilnya
                validation_results = validate_input_output(input_file_path, all_sheets_data, process_type)
                
                # Tulis header untuk sheet validasi
                validation_headers = ["Sheet", "Kolom Input", "Kolom Output", "Status", "Jumlah Data Input", "Jumlah Data Output", "Jumlah Data Sama", "Persentase Kecocokan"]
                for col_idx, header in enumerate(validation_headers, start=1):
                    cell = validation_ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                # Tulis hasil validasi
                for row_idx, result in enumerate(validation_results, start=2):
                    for col_idx, value in enumerate(result, start=1):
                        cell = validation_ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Format khusus untuk baris detail
                        if col_idx == 1 and value.startswith("Detail Perbedaan"):
                            cell.font = Font(bold=True, color="0000FF")  # Biru
                            # Gabungkan sel untuk header detail
                            validation_ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
                        
                        # Format khusus untuk data yang hanya ada di input/output
                        if col_idx == 1 and (value == "Data hanya di Input:" or value == "Data hanya di Output:"):
                            cell.font = Font(italic=True, color="800000")  # Merah tua
                        
                        # Beri warna pada status
                        if col_idx == 4:  # Kolom Status
                            if value == "OK":
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Hijau
                                cell.font = Font(color="006100")  # Hijau tua
                            elif value == "Perlu Dicek":
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Kuning
                                cell.font = Font(color="9C5700")  # Oranye tua
                            elif value == "Tidak Cocok":
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Merah
                                cell.font = Font(color="9C0006")  # Merah tua
                
                # Atur lebar kolom
                for col_idx in range(1, len(validation_headers) + 1):
                    validation_ws.column_dimensions[get_column_letter(col_idx)].width = 20
                
                # Tambahkan filter
                validation_ws.auto_filter.ref = f"A1:{get_column_letter(len(validation_headers))}{len(validation_results) + 1}"
                
                # Tambahkan instruksi penggunaan di atas tabel
                instruction_row = validation_ws.max_row + 2
                cell = validation_ws.cell(row=instruction_row, column=1, value="Petunjuk Penggunaan Sheet Validasi:")
                cell.font = Font(bold=True, size=12)
                
                instructions = [
                    "1. Sheet ini menampilkan perbandingan antara data di file input dan output.",
                    "2. Status 'OK' berarti data cocok 100%.",
                    "3. Status 'Perlu Dicek' berarti kecocokan antara 90% - 99%.",
                    "4. Status 'Tidak Cocok' berarti kecocokan kurang dari 90%.",
                    "5. Detail perbedaan menampilkan contoh data yang hanya ada di input atau output (maksimal 5 contoh).",
                    "6. Lihat sheet 'Analisa Validasi' untuk visualisasi grafik dari hasil validasi."
                ]
                
                for i, instruction in enumerate(instructions):
                    cell = validation_ws.cell(row=instruction_row + i + 1, column=1, value=instruction)
                    # Gabungkan sel untuk instruksi
                    validation_ws.merge_cells(start_row=instruction_row + i + 1, start_column=1, end_row=instruction_row + i + 1, end_column=8)
                
                # Buat chart analisa dan statistik
                create_validation_charts(wb, validation_results)
                
                logging.info(f"Validasi berhasil ditambahkan ke file output")
            except Exception as e:
                logging.error(f"Gagal melakukan validasi: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
        else:
            if not input_file_path:
                logging.warning("Input file path tidak ditemukan untuk validasi")
            else:
                logging.warning(f"File input tidak ditemukan di path: {input_file_path}")
        
        # Simpan workbook
        wb.save(output_path)
        
        # Hitung jumlah sheet dan total record untuk riwayat
        sheet_count = len(all_sheets_data)
        record_count = 0
        for df_item in all_sheets_data.values():
            if isinstance(df_item, dict):
                record_count += sum(len(df) for df in df_item.values())
            else:
                record_count += len(df_item)
        
        # Simpan ke riwayat
        try:
            input_file = os.path.basename(app.config.get('last_input_file', 'Unknown'))
            db_helper.save_process_history(
                file_name=input_file,
                process_type=process_type,
                client=client_name,
                created_by=created_by,
                output_file=output_file_name,
                sheet_count=sheet_count,
                record_count=record_count
            )
            logging.info(f"Riwayat pemrosesan berhasil dicatat: {output_file_name}")
        except Exception as e:
            logging.error(f"Gagal mencatat riwayat pemrosesan: {str(e)}")
        
        return output_file_name
        
    except Exception as e:
        logging.error(f"Error in save_output: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        raise

# Fungsi untuk memvalidasi data antara file input dan output
def validate_input_output(input_file_path, all_sheets_data, process_type):
    try:
        validation_results = []
        
        # Baca file input
        input_excel = pd.ExcelFile(input_file_path)
        logging.info(f"Membaca file input untuk validasi: {input_file_path}")
        logging.info(f"Sheet yang ditemukan di file input: {input_excel.sheet_names}")
        
        # Tentukan kolom yang akan divalidasi berdasarkan jenis proses
        validation_columns = {}
        if process_type == "Bundle":
            validation_columns = {
                "SKU": ["MainSKU", "Main SKU", "SKU"],
                "Qty": ["Qty", "QTY", "Quantity"],
                "Client": ["Client", "CLIENT", "Klien"]
            }
            logging.info(f"Menggunakan kolom validasi untuk proses Bundle: {validation_columns}")
        elif process_type == "Supplementary":
            validation_columns = {
                "Item ID": ["ItemID", "Item ID", "ID"],
                "Client": ["Client", "CLIENT", "Klien"]
            }
            logging.info(f"Menggunakan kolom validasi untuk proses Supplementary: {validation_columns}")
        elif process_type == "Gift":
            validation_columns = {
                "SKU": ["SKU", "Main SKU", "MainSKU"],
                "Client": ["Client", "CLIENT", "Klien"],
                "Gift SKU": ["GiftSKU", "Gift SKU", "GIFT SKU"]
            }
            logging.info(f"Menggunakan kolom validasi untuk proses Gift: {validation_columns}")
        
        # Tambahkan kolom date dan bomSKU untuk validasi baru
        date_columns = {
            "Start Date": ["start_date", "StartDate", "Start Date", "STARTDATE"],
            "End Date": ["end_date", "EndDate", "End Date", "ENDDATE"]
        }
        
        bom_columns = {
            "SKU Product": ["SKU", "MainSKU", "Main SKU", "SKU Product"],
            "Marketplace": ["Marketplace", "MARKETPLACE", "Platform"],
            "BOM SKU": ["bomSKU", "BomSKU", "BOM SKU", "BOMSKU"]
        }
        
        # Lakukan validasi untuk setiap sheet
        for sheet_name in input_excel.sheet_names:
            # Baca sheet input
            try:
                input_df = pd.read_excel(input_excel, sheet_name=sheet_name)
                logging.info(f"Berhasil membaca sheet input: {sheet_name}")
                logging.info(f"Kolom yang ditemukan di sheet input: {input_df.columns.tolist()}")
            except Exception as e:
                logging.error(f"Gagal membaca sheet input {sheet_name}: {str(e)}")
                continue
            
            # Cek apakah sheet ini ada di output
            output_df = None
            output_sheet_name = ""
            
            for out_sheet_name, df in all_sheets_data.items():
                if isinstance(df, dict):
                    # Untuk proses Gift yang memiliki sub-sheet
                    for sub_sheet_name, sub_df in df.items():
                        if sheet_name in out_sheet_name or sheet_name in sub_sheet_name:
                            output_df = sub_df
                            output_sheet_name = f"{out_sheet_name} - {sub_sheet_name}"
                            break
                else:
                    if sheet_name in out_sheet_name:
                        output_df = df
                        output_sheet_name = out_sheet_name
                        break
            
            if output_df is None:
                logging.warning(f"Tidak menemukan sheet output yang sesuai untuk sheet input: {sheet_name}")
                # Coba cari sheet output dengan nama yang mirip
                for out_sheet_name, df in all_sheets_data.items():
                    if isinstance(df, dict):
                        for sub_sheet_name, sub_df in df.items():
                            output_df = sub_df
                            output_sheet_name = f"{out_sheet_name} - {sub_sheet_name}"
                            logging.info(f"Menggunakan sheet output alternatif: {output_sheet_name}")
                            break
                    else:
                        output_df = df
                        output_sheet_name = out_sheet_name
                        logging.info(f"Menggunakan sheet output alternatif: {output_sheet_name}")
                        break
                    if output_df is not None:
                        break
            
            if output_df is None:
                logging.error(f"Tidak dapat menemukan sheet output yang sesuai untuk sheet input: {sheet_name}")
                validation_results.append([
                    sheet_name,
                    "N/A",
                    "N/A",
                    "Tidak Ada Sheet Output",
                    0,
                    0,
                    0,
                    "0%"
                ])
                continue
            
            logging.info(f"Berhasil menemukan sheet output: {output_sheet_name}")
            logging.info(f"Kolom yang ditemukan di sheet output: {output_df.columns.tolist()}")
            
            # Validasi kolom-kolom yang sesuai
            for input_col_key, output_col_options in validation_columns.items():
                # Cari kolom input yang cocok
                input_col = None
                for col in input_df.columns:
                    if input_col_key.lower() in col.lower() or any(key.lower() in col.lower() for key in output_col_options):
                        input_col = col
                        break
                
                if input_col is None:
                    logging.warning(f"Tidak menemukan kolom input yang cocok untuk {input_col_key} di sheet {sheet_name}")
                    continue
                
                # Cari kolom output yang cocok
                output_col = None
                for col in output_df.columns:
                    if any(option.lower() in col.lower() for option in output_col_options):
                        output_col = col
                        break
                
                if output_col is None:
                    logging.warning(f"Tidak menemukan kolom output yang cocok untuk {input_col_key} di sheet {output_sheet_name}")
                    continue
                
                logging.info(f"Validasi kolom: Input '{input_col}' dengan Output '{output_col}'")
                
                # Bersihkan data untuk perbandingan
                input_values = input_df[input_col].astype(str).str.strip().dropna()
                output_values = output_df[output_col].astype(str).str.strip().dropna()
                
                # Hitung jumlah data yang sama
                common_values = set(input_values).intersection(set(output_values))
                common_count = len(common_values)
                
                # Temukan data yang berbeda
                input_only_values = set(input_values) - set(output_values)
                output_only_values = set(output_values) - set(input_values)
                
                # Ambil sampel data yang berbeda (maksimal 5)
                input_only_sample = list(input_only_values)[:5] if input_only_values else []
                output_only_sample = list(output_only_values)[:5] if output_only_values else []
                
                # Hitung persentase kecocokan
                input_count = len(input_values)
                output_count = len(output_values)
                
                logging.info(f"Jumlah data input: {input_count}, output: {output_count}, sama: {common_count}")
                if input_only_sample:
                    logging.info(f"Contoh data yang hanya ada di input: {input_only_sample}")
                if output_only_sample:
                    logging.info(f"Contoh data yang hanya ada di output: {output_only_sample}")
                
                if input_count > 0:
                    match_percentage = (common_count / input_count) * 100
                else:
                    match_percentage = 0
                
                # Tentukan status
                status = "OK"
                if match_percentage < 90:
                    status = "Tidak Cocok"
                elif match_percentage < 100:
                    status = "Perlu Dicek"
                
                # Tambahkan hasil validasi
                validation_results.append([
                    sheet_name,
                    input_col,
                    output_col,
                    status,
                    input_count,
                    output_count,
                    common_count,
                    f"{match_percentage:.2f}%"
                ])
                
                # Jika ada perbedaan, tambahkan detail perbedaan
                if input_only_sample or output_only_sample:
                    # Tambahkan baris kosong untuk pemisah
                    validation_results.append([
                        "", "", "", "", "", "", "", ""
                    ])
                    
                    # Tambahkan header untuk detail perbedaan
                    validation_results.append([
                        f"Detail Perbedaan untuk {sheet_name}: {input_col} vs {output_col}",
                        "", "", "", "", "", "", ""
                    ])
                    
                    # Tambahkan data yang hanya ada di input
                    if input_only_sample:
                        validation_results.append([
                            "Data hanya di Input:",
                            ", ".join(input_only_sample),
                            "", "", "", "", "", ""
                        ])
                    
                    # Tambahkan data yang hanya ada di output
                    if output_only_sample:
                        validation_results.append([
                            "Data hanya di Output:",
                            "", 
                            ", ".join(output_only_sample),
                            "", "", "", "", ""
                        ])
                    
                    # Tambahkan baris kosong untuk pemisah
                    validation_results.append([
                        "", "", "", "", "", "", "", ""
                    ])
            
            # VALIDASI BARU 1: VALIDASI START DATE DAN END DATE
            validation_results.append([
                "", "", "", "", "", "", "", ""
            ])
            validation_results.append([
                f"VALIDASI TANGGAL untuk {sheet_name}",
                "", "", "", "", "", "", ""
            ])
            
            for date_key, date_col_options in date_columns.items():
                # Cari kolom tanggal di input
                input_date_col = None
                for col in input_df.columns:
                    if any(option.lower() in col.lower() for option in date_col_options):
                        input_date_col = col
                        break
                
                # Cari kolom tanggal di output
                output_date_col = None
                for col in output_df.columns:
                    if any(option.lower() in col.lower() for option in date_col_options):
                        output_date_col = col
                        break
                
                if input_date_col is not None and output_date_col is not None:
                    validation_results.append([
                        sheet_name,
                        f"{date_key}",
                        "Perbandingan Input vs Output",
                        "", "", "", "", ""
                    ])
                    
                    # Ambil sampel tanggal (maksimal 10 baris)
                    sample_rows = min(10, len(input_df))
                    for i in range(sample_rows):
                        input_date_val = input_df[input_date_col].iloc[i] if i < len(input_df) else ""
                        output_date_val = output_df[output_date_col].iloc[i] if i < len(output_df) else ""
                        
                        # Tentukan status kecocokan
                        date_match_status = "OK" if str(input_date_val).strip() == str(output_date_val).strip() else "Berbeda"
                        
                        validation_results.append([
                            f"Baris {i+1}",
                            str(input_date_val),
                            str(output_date_val),
                            date_match_status,
                            "", "", "", ""
                        ])
                    
                    validation_results.append([
                        "", "", "", "", "", "", "", ""
                    ])
            
            # VALIDASI BARU 2: VALIDASI SKU PRODUCT × MARKETPLACE VS BOM SKU
            sku_col = None
            marketplace_col = None
            bom_sku_col = None
            
            # Cari kolom SKU product
            for col in input_df.columns:
                if any(sku_option.lower() in col.lower() for sku_option in bom_columns["SKU Product"]):
                    sku_col = col
                    break
            
            # Cari kolom Marketplace
            for col in input_df.columns:
                if any(mp_option.lower() in col.lower() for mp_option in bom_columns["Marketplace"]):
                    marketplace_col = col
                    break
                    
            # Cari kolom BOM SKU
            for col in output_df.columns:
                if any(bom_option.lower() in col.lower() for bom_option in bom_columns["BOM SKU"]):
                    bom_sku_col = col
                    break
            
            if sku_col is not None and marketplace_col is not None and bom_sku_col is not None:
                # Hitung jumlah SKU product unik
                unique_skus = input_df[sku_col].dropna().unique()
                sku_count = len(unique_skus)
                
                # Hitung jumlah marketplace unik
                unique_marketplaces = input_df[marketplace_col].dropna().unique()
                marketplace_count = len(unique_marketplaces)
                
                # Hitung jumlah BOM SKU
                bom_sku_count = len(output_df[bom_sku_col].dropna().unique())
                
                # Hitung perkiraan (SKU × Marketplace)
                expected_bom_count = sku_count * marketplace_count
                
                # Tentukan status perbandingan
                if expected_bom_count == bom_sku_count:
                    bom_status = "OK"
                elif abs(expected_bom_count - bom_sku_count) / expected_bom_count <= 0.1:  # Toleransi 10%
                    bom_status = "Perlu Dicek"
                else:
                    bom_status = "Tidak Cocok"
                
                # Hitung persentase kecocokan
                if expected_bom_count > 0:
                    bom_percentage = (bom_sku_count / expected_bom_count) * 100
                else:
                    bom_percentage = 0
                
                validation_results.append([
                    sheet_name,
                    "Validasi SKU × Marketplace vs BOM SKU",
                    "",
                    bom_status,
                    "", "", "", ""
                ])
                
                validation_results.append([
                    "SKU Product",
                    str(sku_count),
                    "SKU unik",
                    "", "", "", "", ""
                ])
                
                validation_results.append([
                    "Marketplace",
                    str(marketplace_count),
                    "Marketplace unik",
                    "", "", "", "", ""
                ])
                
                validation_results.append([
                    "SKU × Marketplace",
                    str(expected_bom_count),
                    "Hasil perkalian",
                    "", "", "", "", ""
                ])
                
                validation_results.append([
                    "BOM SKU",
                    str(bom_sku_count),
                    "BOM SKU unik",
                    "", "", "", "", ""
                ])
                
                validation_results.append([
                    "Persentase",
                    f"{bom_percentage:.2f}%",
                    "Kecocokan",
                    bom_status,
                    "", "", "", ""
                ])
                
                validation_results.append([
                    "", "", "", "", "", "", "", ""
                ])
        
        if not validation_results:
            logging.warning("Tidak ada hasil validasi yang ditemukan")
            validation_results.append([
                "Tidak Ada Data",
                "N/A",
                "N/A",
                "Tidak Ada Data",
                0,
                0,
                0,
                "0%"
            ])
        
        return validation_results
    except Exception as e:
        logging.error(f"Error dalam validasi input-output: {str(e)}")
        logging.exception(e)
        return [
            ["Error", str(e), "", "Error", 0, 0, 0, "0%"]
        ]

# Fungsi untuk membaca konfigurasi terenkripsi (tidak ada di app.py)
def read_encrypted_config():
    try:
        # Misalkan Anda memiliki file kunci dan file konfigurasi terenkripsi
        key_file = os.path.join(SCRIPT_DIR, "key.key")
        config_file = os.path.join(SCRIPT_DIR, "config.enc")

        # Membaca kunci
        with open(key_file, 'rb') as file:
            key = file.read()

        # Inisialisasi Fernet dengan kunci
        fernet = Fernet(key)

        # Membaca dan mendekripsi file konfigurasi
        with open(config_file, 'rb') as enc_file:
            encrypted_data = enc_file.read()
            decrypted_data = fernet.decrypt(encrypted_data)

        # Mengembalikan data konfigurasi sebagai dictionary
        return json.loads(decrypted_data.decode('utf-8'))
    except Exception as e:
        logging.error(f"Error reading encrypted config: {str(e)}")
        return None

# Fungsi untuk membuat chart analisa dan statistik
def create_validation_charts(wb, validation_results):
    try:
        # Buat sheet untuk chart
        chart_sheet = wb.create_sheet(title="Analisa Validasi")
        
        # Data untuk chart
        status_counts = {"OK": 0, "Perlu Dicek": 0, "Tidak Cocok": 0, "Tidak Ada Data": 0}
        sheet_data = {}
        column_data = {}
        
        # Hitung jumlah status dan data per sheet/kolom
        for result in validation_results:
            if len(result) >= 8 and result[3] in status_counts:
                status_counts[result[3]] += 1
                
                # Hitung data per sheet
                sheet_name = result[0]
                if sheet_name and sheet_name not in ["", "Tidak Ada Data", "Error"]:
                    if sheet_name not in sheet_data:
                        sheet_data[sheet_name] = {"OK": 0, "Perlu Dicek": 0, "Tidak Cocok": 0}
                    sheet_data[sheet_name][result[3]] += 1
                
                # Hitung data per kolom
                if result[1] and result[2] and result[1] not in ["", "N/A", "Data hanya di Input:", "Data hanya di Output:"]:
                    column_key = f"{result[1]} → {result[2]}"
                    if column_key not in column_data:
                        column_data[column_key] = {"OK": 0, "Perlu Dicek": 0, "Tidak Cocok": 0}
                    column_data[column_key][result[3]] += 1
        
        # 1. Buat chart status validasi (pie chart)
        # Tulis data untuk pie chart
        chart_sheet.cell(row=1, column=1, value="Status Validasi").font = Font(bold=True, size=14)
        chart_sheet.cell(row=2, column=1, value="Status")
        chart_sheet.cell(row=2, column=2, value="Jumlah")
        
        row = 3
        for status, count in status_counts.items():
            chart_sheet.cell(row=row, column=1, value=status)
            chart_sheet.cell(row=row, column=2, value=count)
            row += 1
        
        # Buat pie chart
        pie = PieChart()
        pie.title = "Distribusi Status Validasi"
        labels = Reference(chart_sheet, min_col=1, min_row=3, max_row=row-1)
        data = Reference(chart_sheet, min_col=2, min_row=2, max_row=row-1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Tambahkan warna sesuai status
        slice_colors = [
            "C6EFCE",  # Hijau untuk OK
            "FFEB9C",  # Kuning untuk Perlu Dicek
            "FFC7CE",  # Merah untuk Tidak Cocok
            "D9D9D9"   # Abu-abu untuk Tidak Ada Data
        ]
        
        for i, slice in enumerate(pie.series[0].points):
            slice.graphicalProperties.solidFill = slice_colors[i]
        
        # Tambahkan pie chart ke sheet
        chart_sheet.add_chart(pie, "A10")
        
        # 2. Buat chart perbandingan per sheet (bar chart)
        if sheet_data:
            chart_sheet.cell(row=1, column=5, value="Perbandingan per Sheet").font = Font(bold=True, size=14)
            chart_sheet.cell(row=2, column=5, value="Sheet")
            chart_sheet.cell(row=2, column=6, value="OK")
            chart_sheet.cell(row=2, column=7, value="Perlu Dicek")
            chart_sheet.cell(row=2, column=8, value="Tidak Cocok")
            
            row = 3
            for sheet_name, counts in sheet_data.items():
                chart_sheet.cell(row=row, column=5, value=sheet_name)
                chart_sheet.cell(row=row, column=6, value=counts["OK"])
                chart_sheet.cell(row=row, column=7, value=counts["Perlu Dicek"])
                chart_sheet.cell(row=row, column=8, value=counts["Tidak Cocok"])
                row += 1
            
            # Buat bar chart
            bar = BarChart()
            bar.type = "col"
            bar.style = 10
            bar.title = "Status Validasi per Sheet"
            bar.y_axis.title = "Jumlah"
            bar.x_axis.title = "Sheet"
            
            data = Reference(chart_sheet, min_col=6, min_row=2, max_row=row-1, max_col=8)
            cats = Reference(chart_sheet, min_col=5, min_row=3, max_row=row-1)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            
            # Tambahkan warna sesuai status
            bar.series[0].graphicalProperties.solidFill = "C6EFCE"  # Hijau untuk OK
            bar.series[1].graphicalProperties.solidFill = "FFEB9C"  # Kuning untuk Perlu Dicek
            bar.series[2].graphicalProperties.solidFill = "FFC7CE"  # Merah untuk Tidak Cocok
            
            # Tambahkan bar chart ke sheet
            chart_sheet.add_chart(bar, "E10")
        
        # 3. Buat chart persentase kecocokan (column chart)
        # Ambil data persentase kecocokan
        match_data = []
        for result in validation_results:
            if len(result) >= 8 and result[0] not in ["", "Tidak Ada Data", "Error"] and result[7] != "":
                try:
                    # Ekstrak persentase dari string (misalnya "95.50%")
                    percentage_str = result[7].replace("%", "")
                    percentage = float(percentage_str)
                    match_data.append((result[0], result[1], result[2], percentage))
                except (ValueError, TypeError):
                    continue
        
        if match_data:
            chart_sheet.cell(row=30, column=1, value="Persentase Kecocokan").font = Font(bold=True, size=14)
            chart_sheet.cell(row=31, column=1, value="Sheet")
            chart_sheet.cell(row=31, column=2, value="Kolom Input")
            chart_sheet.cell(row=31, column=3, value="Kolom Output")
            chart_sheet.cell(row=31, column=4, value="Persentase")
            
            row = 32
            for item in match_data:
                chart_sheet.cell(row=row, column=1, value=item[0])
                chart_sheet.cell(row=row, column=2, value=item[1])
                chart_sheet.cell(row=row, column=3, value=item[2])
                chart_sheet.cell(row=row, column=4, value=item[3])
                
                # Tambahkan conditional formatting
                cell = chart_sheet.cell(row=row, column=4)
                if item[3] < 90:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif item[3] < 100:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                row += 1
            
            # Buat column chart
            col = BarChart()
            col.type = "col"
            col.style = 10
            col.title = "Persentase Kecocokan per Kolom"
            col.y_axis.title = "Persentase (%)"
            col.x_axis.title = "Kolom"
            
            data = Reference(chart_sheet, min_col=4, min_row=31, max_row=row-1)
            cats = Reference(chart_sheet, min_col=1, min_row=32, max_row=row-1)
            col.add_data(data, titles_from_data=True)
            col.set_categories(cats)
            
            # Tambahkan column chart ke sheet
            chart_sheet.add_chart(col, "A40")
        
        # Atur lebar kolom
        for col in range(1, 10):
            chart_sheet.column_dimensions[get_column_letter(col)].width = 20
        
        # Tambahkan ringkasan statistik
        chart_sheet.cell(row=row + 5, column=1, value="Ringkasan Statistik").font = Font(bold=True, size=14)
        
        total_validations = sum(status_counts.values())
        success_rate = (status_counts["OK"] / total_validations * 100) if total_validations > 0 else 0
        
        chart_sheet.cell(row=row + 6, column=1, value="Total Validasi:")
        chart_sheet.cell(row=row + 6, column=2, value=total_validations)
        
        chart_sheet.cell(row=row + 7, column=1, value="Tingkat Keberhasilan:")
        success_cell = chart_sheet.cell(row=row + 7, column=2, value=f"{success_rate:.2f}%")
        
        # Beri warna pada tingkat keberhasilan
        if success_rate < 90:
            success_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif success_rate < 100:
            success_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            success_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        logging.info(f"Chart analisa validasi berhasil dibuat")
        return True
    except Exception as e:
        logging.error(f"Error in create_validation_charts: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return False

@app.route('/api/validation_data/<filename>', methods=['GET'])
@login_required
def get_validation_data(filename):
    try:
        # Pastikan filename aman
        if not allowed_file(filename):
            return jsonify({'error': 'Format file tidak didukung'}), 400
        
        # Buat path file
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        
        # Cek apakah file ada
        if not os.path.exists(file_path):
            return jsonify({'error': 'File tidak ditemukan'}), 404
        
        # Baca file Excel
        validation_data = {
            'status_counts': {'OK': 0, 'Perlu Dicek': 0, 'Tidak Cocok': 0, 'Tidak Ada Data': 0},
            'sheet_data': {},
            'match_percentages': [],
            'date_validation': [],
            'sku_marketplace_validation': []
        }
        
        try:
            # Buka workbook
            wb = load_workbook(file_path, read_only=True)
            
            # Cek apakah sheet Validasi ada
            if 'Validasi' not in wb.sheetnames:
                return jsonify({'error': 'Sheet Validasi tidak ditemukan dalam file'}), 404
            
            # Baca sheet Validasi
            ws = wb['Validasi']
            
            # Baca data validasi
            headers = []
            validation_results = []
            
            for row_idx, row in enumerate(ws.rows):
                if row_idx == 0:
                    # Baca header
                    headers = [cell.value for cell in row]
                else:
                    # Baca data
                    row_data = [cell.value for cell in row]
                    if len(row_data) >= 8:
                        validation_results.append(row_data)
            
            # Indikator untuk mengidentifikasi bagian-bagian validasi
            in_date_section = False
            in_sku_marketplace_section = False
            current_sheet = ""
            current_date_type = ""
            date_comparison_data = {}
            
            # Hitung jumlah status dan data per sheet/kolom
            for result in validation_results:
                # Hanya proses baris yang memiliki status
                if result[0] and len(result) >= 4 and result[3] in validation_data['status_counts']:
                    validation_data['status_counts'][result[3]] += 1
                    
                    # Hitung data per sheet
                    sheet_name = result[0]
                    if sheet_name and sheet_name not in ["", "Tidak Ada Data", "Error", "Detail Perbedaan", "Data hanya di Input:", "Data hanya di Output:"]:
                        if sheet_name not in validation_data['sheet_data']:
                            validation_data['sheet_data'][sheet_name] = {'OK': 0, 'Perlu Dicek': 0, 'Tidak Cocok': 0}
                        validation_data['sheet_data'][sheet_name][result[3]] += 1
                    
                    # Ambil persentase kecocokan
                    if result[7] and "%" in str(result[7]):
                        try:
                            percentage_str = str(result[7]).replace("%", "")
                            percentage = float(percentage_str)
                            validation_data['match_percentages'].append({
                                'sheet': result[0],
                                'input_column': result[1],
                                'output_column': result[2],
                                'percentage': percentage
                            })
                        except (ValueError, TypeError):
                            pass
                
                # Cek apakah ini adalah bagian validasi tanggal
                if result[0] and "VALIDASI TANGGAL" in str(result[0]):
                    in_date_section = True
                    in_sku_marketplace_section = False
                    current_sheet = result[0].replace("VALIDASI TANGGAL untuk ", "")
                    continue
                
                # Cek apakah ini adalah bagian validasi SKU × Marketplace
                if result[1] and "Validasi SKU × Marketplace vs BOM SKU" in str(result[1]):
                    in_date_section = False
                    in_sku_marketplace_section = True
                    current_sheet = result[0]
                    
                    # Inisialisasi data validasi SKU × Marketplace
                    sku_mp_data = {
                        'sheet': current_sheet,
                        'sku_count': 0,
                        'marketplace_count': 0,
                        'expected_bom_count': 0,
                        'actual_bom_count': 0,
                        'percentage': 0,
                        'status': result[3]
                    }
                    validation_data['sku_marketplace_validation'].append(sku_mp_data)
                    continue
                
                # Proses data dalam bagian validasi tanggal
                if in_date_section:
                    if result[1] and result[1] in ["Start Date", "End Date"]:
                        current_date_type = result[1]
                        date_comparison_data = {
                            'sheet': current_sheet,
                            'date_type': current_date_type,
                            'comparisons': []
                        }
                        validation_data['date_validation'].append(date_comparison_data)
                    elif result[0] and result[0].startswith("Baris ") and current_date_type:
                        # Tambahkan perbandingan tanggal
                        for date_data in validation_data['date_validation']:
                            if date_data['sheet'] == current_sheet and date_data['date_type'] == current_date_type:
                                date_data['comparisons'].append({
                                    'row': result[0],
                                    'input_value': result[1],
                                    'output_value': result[2],
                                    'status': result[3]
                                })
                                break
                
                # Proses data dalam bagian validasi SKU × Marketplace
                if in_sku_marketplace_section:
                    if result[0] == "SKU Product" and len(validation_data['sku_marketplace_validation']) > 0:
                        validation_data['sku_marketplace_validation'][-1]['sku_count'] = int(result[1]) if isinstance(result[1], (int, float)) or (isinstance(result[1], str) and result[1].isdigit()) else 0
                    elif result[0] == "Marketplace" and len(validation_data['sku_marketplace_validation']) > 0:
                        validation_data['sku_marketplace_validation'][-1]['marketplace_count'] = int(result[1]) if isinstance(result[1], (int, float)) or (isinstance(result[1], str) and result[1].isdigit()) else 0
                    elif result[0] == "SKU × Marketplace" and len(validation_data['sku_marketplace_validation']) > 0:
                        validation_data['sku_marketplace_validation'][-1]['expected_bom_count'] = int(result[1]) if isinstance(result[1], (int, float)) or (isinstance(result[1], str) and result[1].isdigit()) else 0
                    elif result[0] == "BOM SKU" and len(validation_data['sku_marketplace_validation']) > 0:
                        validation_data['sku_marketplace_validation'][-1]['actual_bom_count'] = int(result[1]) if isinstance(result[1], (int, float)) or (isinstance(result[1], str) and result[1].isdigit()) else 0
                    elif result[0] == "Persentase" and len(validation_data['sku_marketplace_validation']) > 0:
                        try:
                            percentage_str = str(result[1]).replace("%", "")
                            validation_data['sku_marketplace_validation'][-1]['percentage'] = float(percentage_str)
                            validation_data['sku_marketplace_validation'][-1]['status'] = result[3]
                        except (ValueError, TypeError):
                            pass
            
            return jsonify(validation_data)
        except Exception as e:
            logging.error(f"Error saat membaca file validasi: {str(e)}")
            logging.exception(e)
            return jsonify({'error': f'Gagal membaca file validasi: {str(e)}'}), 500
    except Exception as e:
        logging.error(f"Error api validasi: {str(e)}")
        logging.exception(e)
        return jsonify({'error': f'Terjadi kesalahan: {str(e)}'}), 500

@app.route('/api/output_data/<filename>', methods=['GET'])
@login_required
def get_output_data(filename):
    try:
        # Pastikan filename aman
        if not allowed_file(filename):
            return jsonify({'error': 'Format file tidak didukung'}), 400
        
        # Buat path file
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        
        # Cek apakah file ada
        if not os.path.exists(file_path):
            return jsonify({'error': 'File tidak ditemukan'}), 404
        
        # Baca file Excel
        output_data = {
            'sheets': [],
            'summary': {
                'total_records': 0,
                'sheet_count': 0,
                'columns_per_sheet': {}
            },
            'data_samples': {}
        }
        
        try:
            # Buka workbook
            wb = load_workbook(file_path, read_only=True)
            
            # Kumpulkan data dari semua sheet (kecuali sheet validasi dan analisa)
            excluded_sheets = ['Validasi', 'Analisa Validasi']
            valid_sheets = [sheet for sheet in wb.sheetnames if sheet not in excluded_sheets and not sheet.startswith('Input_')]
            
            output_data['summary']['sheet_count'] = len(valid_sheets)
            
            for sheet_name in valid_sheets:
                ws = wb[sheet_name]
                
                # Ambil header
                headers = []
                for cell in next(ws.rows):
                    headers.append(cell.value)
                
                # Hitung jumlah baris
                row_count = sum(1 for _ in ws.rows) - 1  # Kurangi 1 untuk header
                
                # Tambahkan informasi sheet
                sheet_info = {
                    'name': sheet_name,
                    'row_count': row_count,
                    'column_count': len(headers),
                    'headers': headers
                }
                
                output_data['sheets'].append(sheet_info)
                output_data['summary']['total_records'] += row_count
                output_data['summary']['columns_per_sheet'][sheet_name] = len(headers)
                
                # Ambil sampel data (maksimal 10 baris pertama)
                data_sample = []
                for i, row in enumerate(ws.rows):
                    if i == 0:  # Skip header
                        continue
                    if i > 10:  # Maksimal 10 baris
                        break
                    
                    row_data = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_data[headers[j]] = cell.value
                    
                    data_sample.append(row_data)
                
                output_data['data_samples'][sheet_name] = data_sample
            
            # Tambahkan informasi proses
            process_type = "Unknown"
            if valid_sheets:
                first_sheet = wb[valid_sheets[0]]
                headers = [cell.value for cell in next(first_sheet.rows)]
                
                if 'MainSKU' in headers:
                    process_type = "Bundle"
                elif 'ItemID' in headers:
                    process_type = "Supplementary"
                elif 'SKU' in headers and 'GiftSKU' in headers:
                    process_type = "Gift"
            
            output_data['process_type'] = process_type
            
            return jsonify(output_data)
            
        except Exception as e:
            logging.error(f"Error reading output data: {str(e)}")
            return jsonify({'error': f'Error reading output data: {str(e)}'}), 500
            
    except Exception as e:
        logging.error(f"Error in get_output_data: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

@app.route('/api/clear_temp_files', methods=['POST'])
@login_required
def clear_temp_files():
    """Membersihkan file sementara yang terkait dengan sesi pengguna saat ini."""
    try:
        user = session.get('username', 'unknown')
        logging.info(f"Membersihkan file sementara untuk pengguna: {user}")
        
        # Hapus file di folder uploads yang dibuat oleh pengguna ini
        temp_files_removed = 0
        
        # Periksa folder upload
        for filename in os.listdir(UPLOAD_FOLDER):
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            # Hanya hapus file (bukan folder) yang dimiliki oleh pengguna ini
            # Format nama file biasanya mengandung username atau timestamp
            if os.path.isfile(file_path) and (user in filename or 'temp_' in filename):
                try:
                    os.remove(file_path)
                    temp_files_removed += 1
                    logging.info(f"Menghapus file upload: {filename}")
                except Exception as e:
                    logging.error(f"Gagal menghapus file upload {filename}: {str(e)}")
        
        # Periksa folder output
        output_files_removed = 0
        for filename in os.listdir(OUTPUT_FOLDER):
            file_path = os.path.join(OUTPUT_FOLDER, filename)
            # Hanya hapus file output yang terkait dengan pengguna ini
            if os.path.isfile(file_path) and (user in filename or 'temp_' in filename):
                try:
                    os.remove(file_path)
                    output_files_removed += 1
                    logging.info(f"Menghapus file output: {filename}")
                except Exception as e:
                    logging.error(f"Gagal menghapus file output {filename}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Berhasil membersihkan {temp_files_removed} file upload dan {output_files_removed} file output.',
            'upload_files_removed': temp_files_removed,
            'output_files_removed': output_files_removed
        })
    except Exception as e:
        logging.error(f"Error saat membersihkan file sementara: {str(e)}")
        return jsonify({'error': f'Error: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5111))
    
    # Buat rahasia aplikasi
    app.secret_key = secrets.token_hex(16)
    
    # Konfigurasi tambahan untuk Flask
    app.config['SESSION_COOKIE_SECURE'] = False
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    
    # Jalankan aplikasi tanpa SSL/HTTPS
    app.run(host='0.0.0.0', port=port, debug=True, ssl_context=None)
